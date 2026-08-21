"""
FY27 monthly-financials extractor for the LHRC dashboard.

Sources (all live in the sibling "FY27" folder):
  1. "1.BRM_Deck_FY 27 - Treating Report basis*.xlsx"  -> sheet "Act Vs Bud"
       Monthly Apr..Mar FY27 blocks, values already in Rs Cr:
         #1 Revenue        Kochi    (OP / IP / Total)
         #2 Revenue        Calicut  (OP / IP / Total)
         #3 OP Visit       Kochi    (New / Revisit / Total)
         #4 OP Visit       Calicut  (New / Revisit / Total)
         #5 IP Discharge / Occupancy / Total Hospital Revenue  (group)
         #6 Payor Mix      Cash / Insurance / Government
         #7 Payor Mix      Corporate / Domestic / International
       IMPORTANT: we take ACTUALS ONLY from this deck. Its budget columns are on
       the "Treating Report basis" (Kochi FY27 plan 480.5 Cr / YTD 152.0 Cr) which
       is NOT the basis the dashboard runs on. Sameeraj confirmed 20-Aug-2026 that
       the governing basis is the AOP at 575 Cr (YTD Apr-Jul 180.0 Cr). Mixing the
       two would overstate achievement by ~16pp.

  2. "MIS LHRC_<Month> <YY> Kochi.xlsb" / "... Calicut.xlsb" -> sheet "12P+L Summ"
       Full P&L to EBITDA for the report month + FY27 YTD, actual AND budget.
       This pack's budget IS on the AOP basis (YTD 180.0 Cr), so both actual and
       budget are safe to show here.

Deliberately NOT used (verified empty 20-Aug-2026):
  - "13Monthly P&L "  : every line x every month evaluates to 0. Confirmed a real
                        zero (not a stale formula cache) by recalculating through
                        LibreOffice; columns are also dated Apr-25..Mar-26 (FY26).
  - "14KPI"           : raw volume counts are real but FY26 columns; every
                        revenue-derived KPI / realization / ARPOB row is 0.
  So there is no FY27 monthly EBITDA series and no usable ARPOB/ALOS to chart.

Values: BRM sheet is in Rs Cr -> converted to absolute INR here so the whole
dashboard stays in one unit. 12P+L Summ is already absolute INR.
"""

import os, re, json, glob

CR = 1e7

MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep",
          "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# Act Vs Bud section anchors: (header row, key, unit, labels, money-flag per label)
# Each block is 16 rows: header, group labels, col labels, 12 months, YTD.
# money=True -> value is in Rs Cr in the sheet and gets scaled to absolute INR.
# money=False -> a count (visits, discharges) or a ratio (occupancy); left as-is.
BRM_BLOCKS = [
    (5,   "revenue",  "Kochi",   ("op", "ip", "total"),        (True, True, True)),
    (22,  "revenue",  "Calicut", ("op", "ip", "total"),        (True, True, True)),
    (39,  "opVisits", "Kochi",   ("new", "revisit", "total"),  (False, False, False)),
    (56,  "opVisits", "Calicut", ("new", "revisit", "total"),  (False, False, False)),
    (73,  "hospital", None,      ("ipDisch", "occupancy", "totalRev"), (False, False, True)),
    (90,  "payorA",   None,      ("cash", "insurance", "government"), (True, True, True)),
    (107, "payorB",   None,      ("corporate", "domestic", "international"), (True, True, True)),
]

# within a block, the three groups start at these 1-based columns; Act is +0, Bud +1
GROUP_COLS = (3, 7, 11)


def _n(v):
    return float(v) if isinstance(v, (int, float)) else None


def parse_brm(path):
    """Monthly FY27 ACTUALS by unit from the BRM deck's 'Act Vs Bud' sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Act Vs Bud" not in wb.sheetnames:
        return None
    ws = wb["Act Vs Bud"]
    grid = {i: r for i, r in
            enumerate(ws.iter_rows(min_row=1, max_row=130, max_col=15, values_only=True), 1)}

    def cell(row, col):
        r = grid.get(row)
        if not r or col - 1 >= len(r):
            return None
        return _n(r[col - 1])

    out = {}
    for start, key, unit, labels, money_flags in BRM_BLOCKS:
        block = {}
        for gi, lab in enumerate(labels):
            col = GROUP_COLS[gi]
            is_money = money_flags[gi]
            series = []
            for mi in range(12):
                v = cell(start + 3 + mi, col)
                if v is not None and is_money:
                    v = v * CR
                series.append(v)
            y = cell(start + 15, col)
            if y is not None and is_money:
                y = y * CR
            block[lab] = {"monthly": series, "ytd": y, "money": bool(is_money)}
        slot = key if unit is None else f"{key}:{unit}"
        out[slot] = block
    wb.close()
    return out


# ---- 12P+L Summ (per unit) -------------------------------------------------
# Row anchors are stable across both units. col 7 = month actual, 9 = month
# budget, 14 = YTD actual, 16 = YTD budget (1-based).
PL_ROWS = [
    (10, "Operating Revenue",          "revenue"),
    (16, "Direct Costs - Consumables", "consumables"),
    (19, "Net Revenue / Contribution", "contribution"),
    (23, "Staff Costs",                "staff"),
    (25, "Overheads",                  "overheads"),
    (27, "Provision for Bad Debts",    "badDebts"),
    (29, "Total Expenses",             "totalExp"),
    (31, "Operating Profit (EBITDA)",  "ebitda"),
    (35, "Net EBITDA",                 "netEbitda"),
    (37, "Finance Charges",            "finance"),
    (39, "Cash Profit",                "cashProfit"),
    (41, "Depreciation",               "depreciation"),
    (43, "Net Profit",                 "netProfit"),
]
PL_COLS = {"monthAct": 7, "monthBud": 9, "ytdAct": 14, "ytdBud": 16}


def parse_pl_summ(path):
    """P&L to EBITDA for the report month + FY27 YTD from an .xlsb MIS pack."""
    from pyxlsb import open_workbook
    wb = open_workbook(path)
    target = None
    for s in wb.sheets:
        if s.strip().lower().startswith("12p+l"):
            target = s
            break
    if not target:
        return None

    grid = {}
    with wb.get_sheet(target) as sh:
        for i, row in enumerate(sh.rows(), 1):
            if i > 60:
                break
            grid[i] = {c.c + 1: c.v for c in row}

    def cell(row, col):
        return _n(grid.get(row, {}).get(col))

    lines = []
    for rownum, label, key in PL_ROWS:
        rec = {"key": key, "label": label}
        got = False
        for name, col in PL_COLS.items():
            v = cell(rownum, col)
            rec[name] = v
            if v not in (None, 0):
                got = True
        if got:
            lines.append(rec)

    # month header (row 6, col 7) is an Excel serial for the report month
    mser = cell(6, 7)
    month = None
    if mser:
        import datetime as _dt
        try:
            month = (_dt.date(1899, 12, 30) + _dt.timedelta(days=int(mser))).strftime("%b-%y")
        except Exception:
            month = None
    return {"month": month, "lines": lines}


# ---- AOP performance tracker ---------------------------------------------
# "VPS Lakehsore FY 27 Dashboard_<Mon>'YY.xlsx" — the AOP tracker. Three sheets:
#   Performance Tracker : per-P&L-line FY26 actual, FY27 plan, YTD plan/actual/%achv,
#                         then 12 month blocks of 4 cols (Plan, Actual, %Achv, %Rev)
#                         starting at col 12 (Apr) and stepping 4.
#   Dashboard           : chart block rows 42-53 = monthly revenue & EBITDA plan/actual.
#   Monthly P&L         : rows 5-49 the plan P&L; rows 55-97 the initiative build-up
#                         (col1 name, col2 type, col3 full-year, cols 4-15 monthly,
#                          col16 total, col17 code), totals at rows 98-100.
# All figures are Rs LAKHS in the sheet -> scaled to absolute INR here.
# This workbook is the same plan as the AOP: FY27 total revenue ties to Rs 575.21 Cr
# exactly, and plan-total less IP+OP equals F&B + Other Income to the rupee.
LAKH = 1e5

TRACKER_LINES = [
    "IP Revenue", "IP Pharmacy", "OP Revenue", "OP Pharmacy", "Ayurveda",
    "F&B / VPS Gourmet", "Other Income", "Total Revenue",
    "Consumption / Drug & Consumables", "Doctor Cost", "Total Variable Cost",
    "Contribution",
    "Employee Cost", "Utilities & Power", "Lab Test Charges", "Rent Paid",
    "Repair & Maintenance", "Housekeeping", "Security", "Printing & Stationery",
    "Insurance", "Rates & Taxes", "Professional Fee", "Quality & Infection",
    "Bad Debts Provision", "Marketing", "Communication", "Business Travel",
    "Miscellaneous", "CSR", "Total Fixed Cost",
    "EBITDA", "(-) Depreciation", "(-) Finance Cost", "PBT",
    "(-) Tax @ 25.89%", "PAT",
]
SUBTOTAL_LINES = {"Total Revenue", "Total Variable Cost", "Contribution",
                  "Total Fixed Cost", "EBITDA", "PBT", "PAT"}
COST_LINES = {"Consumption / Drug & Consumables", "Doctor Cost", "Total Variable Cost",
              "Employee Cost", "Utilities & Power", "Lab Test Charges", "Rent Paid",
              "Repair & Maintenance", "Housekeeping", "Security",
              "Printing & Stationery", "Insurance", "Rates & Taxes",
              "Professional Fee", "Quality & Infection", "Bad Debts Provision",
              "Marketing", "Communication", "Business Travel", "Miscellaneous",
              "CSR", "Total Fixed Cost", "(-) Depreciation", "(-) Finance Cost",
              "(-) Tax @ 25.89%"}


def parse_tracker(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {"source": os.path.basename(path)}

    def sc(v):
        v = _n(v)
        return None if v is None else v * LAKH

    # --- Performance Tracker ---
    if "Performance Tracker" in wb.sheetnames:
        ws = wb["Performance Tracker"]
        rows = list(ws.iter_rows(min_row=1, max_row=60, max_col=60, values_only=True))
        byname = {}
        for r in rows:
            if r and isinstance(r[0], str) and r[0].strip() in TRACKER_LINES:
                byname.setdefault(r[0].strip(), r)
        lines = []
        for name in TRACKER_LINES:
            r = byname.get(name)
            if not r:
                continue
            months = []
            for k in range(12):
                c = 12 + 4 * k                       # 1-based col of this month's Plan
                months.append({"plan": sc(r[c - 1]), "act": sc(r[c])})
            lines.append({
                "label": name,
                "isSub": name in SUBTOTAL_LINES,
                "isCost": name in COST_LINES,
                "fy26Act": sc(r[1]),
                "fy27Plan": sc(r[3]),
                "ytdPlan": sc(r[6]),
                "ytdAct": sc(r[7]),
                "months": months,
            })
        out["lines"] = lines

    # --- Dashboard chart block (monthly revenue / EBITDA plan vs actual) ---
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        rev, eb, labels = [], [], []
        for r in ws.iter_rows(min_row=42, max_row=53, max_col=7, values_only=True):
            if not r or not r[0]:
                continue
            labels.append(r[0])
            rev.append({"plan": sc(r[1]), "act": sc(r[2])})
            eb.append({"plan": sc(r[5]), "act": sc(r[6])})
        out["chart"] = {"labels": labels, "revenue": rev, "ebitda": eb}

    # --- initiative build-up ---
    if "Monthly P&L" in wb.sheetnames:
        ws = wb["Monthly P&L"]
        rows = {i: r for i, r in enumerate(
            ws.iter_rows(min_row=1, max_row=115, max_col=17, values_only=True), 1)}
        inits, section = [], None
        for i in range(54, 98):
            r = rows.get(i)
            if not r or not isinstance(r[0], str):
                continue
            name = r[0].strip()
            code = r[16] if len(r) > 16 else None
            typ = r[1]
            if not typ and not code:
                section = name           # a section banner row
                continue
            inits.append({
                "code": (str(code).strip() if code else None),
                "name": name,
                "type": typ,
                "section": section,
                "fullYear": sc(r[2]),
                "months": [sc(r[3 + k]) for k in range(12)],
            })
        out["initiatives"] = inits
        tot = {}
        for i, key in ((98, "revImpact"), (99, "costSaving"), (100, "gross")):
            r = rows.get(i)
            if r:
                tot[key] = {"fullYear": sc(r[2]),
                            "months": [sc(r[3 + k]) for k in range(12)]}
        # rows 103-107 = extra cost needed to land the revenue uplift; 109 = total
        extra = []
        for i in range(103, 108):
            r = rows.get(i)
            if r and isinstance(r[0], str) and r[0].strip():
                extra.append({"label": r[0].strip(),
                              "total": sc(r[15]),
                              "months": [sc(r[3 + k]) for k in range(12)]})
        r109 = rows.get(109)
        tot["extraCost"] = {"lines": extra,
                            "total": sc(r109[15]) if r109 else None,
                            "months": [sc(r109[3 + k]) for k in range(12)] if r109 else []}
        out["initTotals"] = tot

    wb.close()
    return out


# ---- orchestration --------------------------------------------------------

def find_fy27_folder(mis_folder):
    """FY27 sits alongside the Daily MIS Reports folder."""
    parent = os.path.dirname(os.path.abspath(mis_folder))
    for cand in (os.path.join(parent, "FY27"),
                 os.path.join(os.path.dirname(parent), "FY27")):
        if os.path.isdir(cand):
            return cand
    return None


def _sig(p):
    st = os.stat(p)
    return "%s|%d" % (os.path.basename(p), int(st.st_mtime))


def collect(mis_folder, tools_dir, verbose=True):
    """Returns the fy27 dict for the dashboard, or None. Caches by filename|mtime."""
    fy = find_fy27_folder(mis_folder)
    if not fy:
        if verbose:
            print("FY27: folder not found; skipping")
        return None

    cache_path = os.path.join(tools_dir, "fy27_cache.json")
    cache = {}
    if os.path.exists(cache_path):
        try:
            cache = json.load(open(cache_path))
        except Exception:
            cache = {}

    result = {"units": {}, "monthly": {}, "source": {}}

    # --- BRM deck: newest matching file ---
    brms = sorted(glob.glob(os.path.join(fy, "*BRM_Deck_FY 27*Treating Report basis*.xlsx")),
                  key=lambda p: os.stat(p).st_mtime, reverse=True)
    if not brms:
        brms = sorted(glob.glob(os.path.join(fy, "*BRM_Deck_FY 27*.xlsx")),
                      key=lambda p: os.stat(p).st_mtime, reverse=True)
    if brms:
        bp = brms[0]
        k = "brm|" + _sig(bp)
        if k in cache:
            result["monthly"] = cache[k]
        else:
            if verbose:
                print("FY27: parsing BRM", os.path.basename(bp))
            try:
                result["monthly"] = parse_brm(bp) or {}
                cache[k] = result["monthly"]
            except Exception as e:
                print("FY27: BRM parse failed:", e)
        result["source"]["brm"] = os.path.basename(bp)

    # --- MIS packs per unit: newest per unit ---
    for unit in ("Kochi", "Calicut"):
        cands = sorted(glob.glob(os.path.join(fy, "MIS LHRC_*%s.xlsb" % unit)),
                       key=lambda p: os.stat(p).st_mtime, reverse=True)
        if not cands:
            continue
        p = cands[0]
        k = "pl|" + _sig(p)
        if k in cache:
            result["units"][unit] = cache[k]
        else:
            if verbose:
                print("FY27: parsing P&L", os.path.basename(p))
            try:
                pl = parse_pl_summ(p)
                if pl:
                    result["units"][unit] = pl
                    cache[k] = pl
            except Exception as e:
                print("FY27: P&L parse failed for %s: %s" % (unit, e))
        result["source"].setdefault("pl", {})[unit] = os.path.basename(p)

    # --- AOP performance tracker: newest "FY 27 Dashboard_*" workbook ---
    trs = sorted(glob.glob(os.path.join(fy, "*FY 27 Dashboard*.xlsx")),
                 key=lambda p: os.stat(p).st_mtime, reverse=True)
    trs = [t for t in trs if not os.path.basename(t).startswith("~$")]
    if trs:
        tp = trs[0]
        k = "trk|" + _sig(tp)
        if k in cache:
            result["tracker"] = cache[k]
        else:
            if verbose:
                print("FY27: parsing tracker", os.path.basename(tp))
            try:
                t = parse_tracker(tp)
                if t:
                    result["tracker"] = t
                    cache[k] = t
            except Exception as e:
                print("FY27: tracker parse failed:", e)
        result["source"]["tracker"] = os.path.basename(tp)

    try:
        json.dump(cache, open(cache_path, "w"))
    except Exception:
        pass

    if not result["monthly"] and not result["units"] and not result.get("tracker"):
        return None

    # basis note rendered on the page so nobody re-litigates the 152 vs 180 question
    result["basisNote"] = (
        "Monthly figures are ACTUALS from the BRM deck. P&L actual-vs-budget uses the "
        "MIS pack, whose budget ties to the FY27 AOP (Rs 575 Cr plan / Rs 180.0 Cr YTD). "
        "The BRM deck's own 'Treating Report basis' budget (Rs 480.5 Cr plan / Rs 152.0 Cr "
        "YTD) is deliberately not used."
    )
    result["months"] = MONTHS
    return result


if __name__ == "__main__":
    import sys
    folder = sys.argv[1] if len(sys.argv) > 1 else "."
    tools = os.path.join(folder, "_dashboard_tools")
    r = collect(folder, tools)
    print(json.dumps(r, indent=1)[:4000])
