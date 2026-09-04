#!/usr/bin/env python3
"""
VPS Lakeshore (LHRC) — Daily Revenue Dashboard builder (v2, granular).

Scans the "Daily MIS Reports" folder for:
  - Daily Revenue Flash_New_DD-MM-YYYY*.xlsx   (Burjeel flash: revenue, budget,
    collections, dept/doctor day detail, discharge lists)
  - Daily MIS  (DD-MM-YYYY)- VPS Lakeshore*.xlsx (internal master: doctor x day
    revenue / OP visits / admissions / discharges, bed days, ALOS)

and regenerates LHRC_Revenue_Dashboard.html (self-contained).
Doctor/dept month aggregates persist in _dashboard_tools/history.json so past
months survive even if the source files are removed.

Usage:  python3 build_dashboard.py [folder]
"""
import sys, os, re, json, glob, datetime, time
import openpyxl

FY_MONTHS = ["APRIL","MAY","JUNE","JULY","AUGUST","SEPTEMBER","OCTOBER",
             "NOVEMBER","DECEMBER","JANUARY","FEBRUARY","MARCH"]

# ---------------------------------------------------------------------------
# Declared non-working days, other than Sundays. THE ONE LIST TO MAINTAIN.
# The variance panel uses this to separate calendar effect from real
# throughput: everything else on that panel is computed from the data.
# Add movable festival dates (Onam, Vishu, Eid, Easter) as they are declared;
# a month with nothing listed simply reports its Sundays.
# ---------------------------------------------------------------------------
HOLIDAYS = {
    "2026-08-15": "Independence Day",
    "2026-08-26": "Thiruvonam",
    "2026-10-02": "Gandhi Jayanti",
    "2026-12-25": "Christmas",
    "2027-01-26": "Republic Day",
    "2027-05-01": "May Day",
}

def file_date(path):
    m = re.search(r'(\d{2})-(\d{2})-(\d{4})', os.path.basename(path))
    if not m: return None
    d, mo, y = map(int, m.groups())
    try: return datetime.date(y, mo, d)
    except ValueError: return None

def num(v): return float(v) if isinstance(v, (int, float)) else 0.0

def norm(s): return re.sub(r'\s+', ' ', str(s).strip()).upper()

def title(s):
    return ' '.join(w if ('.' in w and len(w) <= 4) else w.capitalize()
                    for w in str(s).strip().split())

# ---------------------------------------------------------------- flash parsers
def parse_year_tables(ws):
    rows = list(ws.iter_rows(values_only=True, max_col=25))
    out, cur = {}, None
    for r in rows:
        a = str(r[0]).strip() if r[0] is not None else ""
        up = a.upper()
        if up.startswith("YEAR 20") or up.startswith("FY 20"):
            cur = re.sub(r'^(YEAR|FY)\s*', 'FY ', a, flags=re.I).strip()
            out[cur] = {"months": []}
        elif cur and up in FY_MONTHS and any(isinstance(x,(int,float)) for x in r[1:7]):
            out[cur]["months"].append(dict(
                month=up.title(), opVisits=num(r[1]), ipDisch=num(r[2]),
                revOP=num(r[3]), revIP=num(r[4]), revPH=num(r[5]), revTot=num(r[6]),
                budOPv=num(r[7]), budIPd=num(r[8]),
                budOP=num(r[9]), budIP=num(r[10]), budPH=num(r[11]), budTot=num(r[12]),
                collCash=num(r[13]), collCredit=num(r[14]), collTot=num(r[15]),
                budCollTot=num(r[18]),
                census=num(r[23]) if len(r) > 24 else 0,
                occDays=num(r[24]) if len(r) > 24 else 0))
    return out

def parse_month_sheet(ws):
    days = []
    for r in ws.iter_rows(values_only=True, max_col=17):
        if isinstance(r[1], (datetime.datetime, datetime.date)):
            days.append(dict(
                date=r[1].strftime("%Y-%m-%d"), dow=str(r[0] or ""),
                opVisits=num(r[2]), ipDisch=num(r[3]),
                revOP=num(r[4]), revIP=num(r[5]), revPH=num(r[6]), revTot=num(r[7]),
                budTot=num(r[13]), collCash=num(r[14]), collCredit=num(r[15]),
                collTot=num(r[16])))
    return days

def parse_dept_sheet(ws):
    header, hrow = None, None
    for i, r in enumerate(ws.iter_rows(values_only=True, max_row=10)):
        vals = [str(v).strip() if v else "" for v in r]
        if "Department" in vals and "Doctor" in vals:
            header, hrow = vals, i + 1; break
    if not header: return {}, {}, {}, {}
    ix = {n: header.index(n) for n in ("Department","Doctor","Revenue","New Type") if n in header}
    if "Revenue" not in ix: return {}, {}, {}, {}
    dept, doc, typ, doctype = {}, {}, {}, {}
    for r in ws.iter_rows(values_only=True, min_row=hrow + 1):
        d = r[ix["Department"]] if ix["Department"] < len(r) else None
        if not d or not str(d).strip(): continue
        rev = num(r[ix["Revenue"]]) if ix["Revenue"] < len(r) else 0.0
        dept[str(d).strip()] = dept.get(str(d).strip(), 0) + rev
        dn, t = None, None
        if "Doctor" in ix and ix["Doctor"] < len(r) and r[ix["Doctor"]]:
            dn = norm(r[ix["Doctor"]]); doc[dn] = doc.get(dn, 0) + rev
        if "New Type" in ix and ix["New Type"] < len(r) and r[ix["New Type"]]:
            t = norm(r[ix["New Type"]])
            if t in ("IP","OP","PH"): typ[t] = typ.get(t, 0) + rev
        if dn and t in ("IP","OP","PH"):
            dd = doctype.setdefault(dn, {"IP":0,"OP":0,"PH":0}); dd[t] += rev
    return dept, doc, typ, doctype

def parse_op_sheet(ws):
    """Flash 'OP' sheet -> {doctor: {new, free, renew, tot}} for that day."""
    header, hrow = None, None
    for i, r in enumerate(ws.iter_rows(values_only=True, max_row=8)):
        vals = [norm(v) if v else "" for v in r]
        if "DOCTOR NAME" in vals and "TOTAL VISIT" in vals:
            header, hrow = vals, i + 1; break
    if not header: return {}
    ix = {n: header.index(n) for n in ("DOCTOR NAME","NEW","FREE","RENEW","TOTAL VISIT") if n in header}
    out = {}
    for r in ws.iter_rows(values_only=True, min_row=hrow + 1):
        dn = r[ix["DOCTOR NAME"]] if ix["DOCTOR NAME"] < len(r) else None
        if not dn or norm(dn) in ("TOTAL","GRAND TOTAL"): continue
        rec = out.setdefault(norm(dn), {"new":0,"free":0,"renew":0,"tot":0})
        rec["new"] += num(r[ix.get("NEW", -1)]) if "NEW" in ix else 0
        rec["free"] += num(r[ix.get("FREE", -1)]) if "FREE" in ix else 0
        rec["renew"] += num(r[ix.get("RENEW", -1)]) if "RENEW" in ix else 0
        rec["tot"] += num(r[ix.get("TOTAL VISIT", -1)]) if "TOTAL VISIT" in ix else 0
    return out

def classify_scheme(s):
    u = norm(s)
    if "CASH" in u: return "Cash"
    if "ECHS" in u: return "ECHS"
    if any(k in u for k in ("INSURANCE","TPA","HEALTH","ASSIST","MEDI","STAR",
                            "NIVA","ICICI","HDFC","BAJAJ","CIGNA","ALLIANZ")):
        return "Insurance"
    return "Corporate/Other"

def classify_status(s):
    u = norm(s)
    if "RECOVER" in u or "IMPROV" in u or "CURED" in u: return "Recovered"
    if "EXPIR" in u or "DEATH" in u or "DECEAS" in u: return "Expired"
    if "DAMA" in u or "LAMA" in u or "AGAINST" in u or "REQUEST" in u: return "DAMA/LAMA"
    if "REFER" in u or "TRANSFER" in u: return "Referred/Transferred"
    return "Other"

def parse_dis_sheet(ws):
    """Flash 'Dis' sheet -> list of discharge dicts."""
    header, hrow = None, None
    for i, r in enumerate(ws.iter_rows(values_only=True, max_row=8)):
        vals = [norm(v) if v else "" for v in r]
        if "DISCHARGE DATE" in vals and "DOCTOR" in vals:
            header, hrow = vals, i + 1; break
    if not header: return []
    ix = {n: header.index(n) for n in
          ("DISCHARGE DATE","ADMISSION DATE","PATIENT NO","DOCTOR","STATUS","SCHEME")
          if n in header}
    out = []
    for r in ws.iter_rows(values_only=True, min_row=hrow + 1):
        dd = r[ix["DISCHARGE DATE"]] if "DISCHARGE DATE" in ix else None
        if not isinstance(dd, (datetime.datetime, datetime.date)): continue
        ad = r[ix.get("ADMISSION DATE", -1)] if "ADMISSION DATE" in ix else None
        los = (dd.date() if isinstance(dd, datetime.datetime) else dd)
        alos = None
        if isinstance(ad, (datetime.datetime, datetime.date)):
            a = ad.date() if isinstance(ad, datetime.datetime) else ad
            alos = max((los - a).days, 0)
        out.append(dict(
            date=los.strftime("%Y-%m-%d"),
            pid=str(r[ix["PATIENT NO"]]) if "PATIENT NO" in ix else "",
            doctor=norm(r[ix["DOCTOR"]]) if "DOCTOR" in ix and r[ix["DOCTOR"]] else "",
            status=classify_status(r[ix["STATUS"]]) if "STATUS" in ix and r[ix["STATUS"]] else "Other",
            scheme=classify_scheme(r[ix["SCHEME"]]) if "SCHEME" in ix and r[ix["SCHEME"]] else "Corporate/Other",
            alos=alos))
    return out

def parse_flash_details(fpath):
    """Parse one flash file's Dept wise / OP / Dis sheets -> JSON-serializable dict."""
    w = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    dept, doc, doctype, op, dis = {}, {}, {}, {}, []
    if "Dept wise" in w.sheetnames:
        dept, doc, _, doctype = parse_dept_sheet(w["Dept wise"])
    if "OP" in w.sheetnames:
        op = parse_op_sheet(w["OP"])
    if "Dis" in w.sheetnames:
        dis = parse_dis_sheet(w["Dis"])
    w.close()
    return {"dept": dept, "doc": doc, "doctype": doctype, "op": op, "dis": dis}

# ------------------------------------------------------------ Daily MIS parsers
def parse_doctor_day_matrix(ws, value_name):
    """Sheets laid out as [Sl, Department, Doctor, <date cols>...] -> {(dept,doc): total}."""
    rows = list(ws.iter_rows(values_only=True))
    hrow, datecols = None, []
    for i, r in enumerate(rows[:8]):
        dc = [(j, c.date() if isinstance(c, datetime.datetime) else c)
              for j, c in enumerate(r) if isinstance(c, (datetime.datetime, datetime.date))]
        if len(dc) >= 5:
            # dominant month only, first occurrence of each date (a trailing
            # 'Total' column reuses day-1's date and would double-count)
            months = {}
            for _, c in dc: months[(c.year, c.month)] = months.get((c.year, c.month), 0) + 1
            dom = max(months, key=months.get)
            seen = set()
            for j, c in dc:
                if (c.year, c.month) == dom and c not in seen:
                    seen.add(c); datecols.append(j)
            hrow = i; break
    if hrow is None: return {}
    # find dept/doctor columns from a header row at/above hrow
    dept_ix, doc_ix = 1, 2
    for r in rows[:hrow + 2]:
        vals = [str(v).strip() if v else "" for v in r]
        if "Department" in vals and "Doctor" in vals:
            dept_ix, doc_ix = vals.index("Department"), vals.index("Doctor")
    day_of = {j: rows[hrow][j].day if isinstance(rows[hrow][j], datetime.date)
              else rows[hrow][j].date().day for j in datecols}
    out, daily, last_dept = {}, {}, ""
    for r in rows[hrow + 1:]:
        dept = str(r[dept_ix]).strip() if len(r) > dept_ix and r[dept_ix] else ""
        if dept: last_dept = dept
        doc = str(r[doc_ix]).strip() if len(r) > doc_ix and r[doc_ix] else ""
        if not doc: continue
        if norm(doc) in ("TOTAL", "GRAND TOTAL"):
            break  # bottom summary block (TOTAL / category rows) — not doctors
        tot = sum(num(r[j]) for j in datecols if j < len(r))
        key = (norm(last_dept), norm(doc))
        out[key] = out.get(key, 0) + tot
        dd = daily.setdefault(norm(doc), {})
        for j in datecols:
            if j < len(r) and num(r[j]):
                dd[day_of[j]] = dd.get(day_of[j], 0) + num(r[j])
    return out, daily

def parse_mom_fy(ws):
    """'MoM FY 26-27' -> {month 'YYYY-MM': {bedCap, occDays, alos}}"""
    rows = list(ws.iter_rows(values_only=True, max_col=16))
    mcols = {}
    for r in rows[:10]:
        dc = [(j, c) for j, c in enumerate(r) if isinstance(c, (datetime.datetime, datetime.date))]
        if len(dc) >= 6:
            mcols = {j: c.strftime("%Y-%m") for j, c in dc}; break
    if not mcols: return {}
    out = {m: {} for m in mcols.values()}
    for r in rows:
        lbl = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        key = None
        if lbl.startswith("Bed Capacity"): key = "bedCap"
        elif lbl.startswith("Bed Occupancy (D"): key = "occDays"
        elif lbl == "ALOS": key = "alos"
        if key:
            for j, m in mcols.items():
                if j < len(r) and isinstance(r[j], (int, float)):
                    out[m][key] = float(r[j])
    return {m: v for m, v in out.items() if v}

# ------------------------------------------------------------------- FY27 AOP
# Operative plan: BRM_Deck/AOP/Final Final AOP/AOP_vMay26_v2.xlsx -> 'Monthly P&L'
# All figures already in Rs Crore. IP+OP is the line comparable to flash gross
# revenue (flash Pharmacy sits inside AOP IP/OP); F&B and Other Income are shown
# separately because the flash excludes them.
AOP_GLOB = "BRM_Deck/AOP/Final Final AOP/AOP_vMay26_v2.xlsx"
AOP_MONTHS = ["April", "May", "June", "July", "August", "September",
              "October", "November", "December", "January", "February", "March"]

def find_aop(folder):
    """Walk up from the MIS folder looking for the AOP workbook.
    AOP_PATH env var short-circuits the walk: set it to an explicit path, or to
    an empty string to skip the search entirely (cached aop_fy27.json is used).
    Needed when the folder is a mount whose parents are huge/unrelated."""
    env = os.environ.get("AOP_PATH")
    if env is not None:
        return env if (env and os.path.exists(env)) else None
    d = os.path.abspath(folder)
    for depth in range(5):
        cand = os.path.join(d, AOP_GLOB)
        if os.path.exists(cand): return cand
        # Recursive glob only inside the MIS folder itself: on sandbox/VM mounts the
        # parents can be enormous (or the whole filesystem) and a recursive walk
        # there takes minutes. Parents get the direct AOP_GLOB check above only.
        if depth == 0:
            hits = glob.glob(os.path.join(d, "**", "AOP_vMay26_v2.xlsx"), recursive=True)
            if hits: return hits[0]
        d = os.path.dirname(d)
    return None

def parse_aop(path):
    """'Monthly P&L' -> {'source','fyPlan','fy26Actual','months':[{month,ip,op,fnb,other,total}]}
    Row labels in col A; Apr-26..Mar-27 in cols D..O (3..14); col 1 = FY26 actual,
    col 2 = FY27 plan. Values are Rs Crore -> stored as absolute INR for the UI."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Monthly P&L"]
    rows = list(ws.iter_rows(values_only=True, max_row=40, max_col=16))
    wb.close()
    want = {"IP Revenue": "ip", "OP Revenue": "op", "F&B / VPS Gourmet": "fnb",
            "Other Income": "other", "Total Revenue": "total"}
    got, fy26, fyplan = {}, {}, {}
    for r in rows:
        lbl = str(r[0]).strip() if r[0] else ""
        if lbl not in want: continue
        k = want[lbl]
        got[k] = [num(r[3 + i]) * 1e7 for i in range(12)]
        fy26[k] = num(r[1]) * 1e7
        fyplan[k] = num(r[2]) * 1e7
    if "total" not in got: return None
    months = [{"month": AOP_MONTHS[i],
               **{k: got.get(k, [0] * 12)[i] for k in ("ip", "op", "fnb", "other", "total")}}
              for i in range(12)]
    for m in months: m["rev"] = m["ip"] + m["op"]          # flash-comparable
    return {"source": os.path.basename(path), "months": months,
            "fyPlan": fyplan, "fy26Actual": fy26,
            "fyPlanRev": sum(m["rev"] for m in months),
            "fyPlanTotal": sum(m["total"] for m in months)}

# ---------------------------------------------------------------------- main
def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else \
        os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    tools = os.path.join(folder, "_dashboard_tools")
    os.makedirs(tools, exist_ok=True)
    hist_path = os.path.join(tools, "history.json")
    history = {}
    if os.path.exists(hist_path):
        try: history = json.load(open(hist_path))
        except Exception: history = {}

    # ---- flash files ----
    by_date = {}
    for f in glob.glob(os.path.join(folder, "**", "Daily Revenue Flash_New_*.xlsx"), recursive=True):
        d = file_date(f)
        if d and (d not in by_date or os.path.getmtime(f) > os.path.getmtime(by_date[d])):
            by_date[d] = f
    if not by_date:
        print("No flash files found in", folder); sys.exit(1)
    dates = sorted(by_date)
    latest_date, latest_file = dates[-1], by_date[dates[-1]]
    print(f"{len(dates)} flash dates; latest {latest_date}")

    wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)
    year_tables = {}
    for sn in wb.sheetnames:
        if re.match(r'YEAR .*-D', sn): year_tables = parse_year_tables(wb[sn])
    summary = {}
    if "Summary_New" in wb.sheetnames:
        rows = list(wb["Summary_New"].iter_rows(values_only=True, max_col=13))
        def stat(lbl):
            for r in rows:
                if r[0] and str(r[0]).strip().startswith(lbl):
                    return [num(x) for x in r[2:5]]
            return [0,0,0]
        summary = {"bedOcc": stat("Bed Occup."), "surgeries": stat("No. of Surgery"),
                   "admissions": stat("Admissions")}
        for r in rows:
            lbl = str(r[6]).strip() if r[6] else ""
            if lbl in ("OP","IP","PH","IP Conv. Rate"):
                summary.setdefault("real", {})[lbl] = [num(r[7]), num(r[8])]
        for r in rows:
            lbl = str(r[4]).strip() if r[4] else ""
            if lbl == "Total No. of Working Days": summary["workingDays"] = num(r[7])
            if lbl == "Current Cumulative days":   summary["cumDays"] = num(r[7])
        # MONTH (Projections) block: ACTUAL / BUDGET rows, Total Revenue col H ('000)
        in_proj = False
        for r in rows:
            a = str(r[0]).strip() if r[0] else ""
            b = str(r[1]).strip() if r[1] else ""
            if a.startswith("MONTH"): in_proj = True; continue
            if in_proj and b == "ACTUAL":  summary["projActual"] = num(r[7]) * 1000
            if in_proj and b == "BUDGET":  summary["monthBudget"] = num(r[7]) * 1000; in_proj = False
    wb.close()

    # daily series: newest flash per month carries the whole month
    month_files = {}
    for d in dates: month_files[(d.year, d.month)] = by_date[d]
    daily = {}
    for (y, mo), f in sorted(month_files.items()):
        w = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sn in w.sheetnames:
            m = re.match(r'([A-Za-z]+)-(\d{4})-D$', sn)
            if m and int(m.group(2)) == y and m.group(1)[:3].upper() in \
                    {mm[:3] for mm in FY_MONTHS} and \
                    FY_MONTHS.index([mm for mm in FY_MONTHS if mm.startswith(m.group(1)[:3].upper())][0]) == (mo - 4) % 12:
                rows = [r for r in parse_month_sheet(w[sn]) if r["revTot"] > 0 or r["budTot"] > 0]
                daily[f"{y}-{mo:02d}"] = rows
        w.close()

    # persist daily grids so months survive after their flash files are deleted
    dgrid_path = os.path.join(tools, "daily_history.json")
    daily_hist = {}
    if os.path.exists(dgrid_path):
        try: daily_hist = json.load(open(dgrid_path))
        except Exception: daily_hist = {}
    for mk, rows in daily.items():
        # keep the richer version (more days) if both exist
        if mk not in daily_hist or len(rows) >= len(daily_hist[mk]):
            daily_hist[mk] = rows
    json.dump(daily_hist, open(dgrid_path, "w"))
    for mk, rows in daily_hist.items():
        if mk not in daily:
            daily[mk] = rows

    # dept/doctor day cuts + discharge lists from every flash
    # (per-file cache keyed by name|mtime so repeat builds skip re-parsing)
    fcache_path = os.path.join(tools, "flash_cache.json")
    fcache = {}
    if os.path.exists(fcache_path):
        try: fcache = json.load(open(fcache_path))
        except Exception: fcache = {}
    dept_tot, doc_tot, dept_dates, discharges, seen_pid = {}, {}, [], [], set()
    doc_type_mix, op_agg = {}, {}
    fcache_dirty = False
    for d in dates:
        fpath = by_date[d]
        ck = f"{os.path.basename(fpath)}|{int(os.path.getmtime(fpath))}"
        c = fcache.get(ck)
        if c is None:
            c = parse_flash_details(fpath)
            fcache[ck] = c; fcache_dirty = True
        if c["dept"]:
            dept_dates.append(d.strftime("%Y-%m-%d"))
            for k, v in c["dept"].items(): dept_tot[k] = dept_tot.get(k, 0) + v
            for k, v in c["doc"].items():  doc_tot[k] = doc_tot.get(k, 0) + v
            for k, v in c["doctype"].items():
                dd = doc_type_mix.setdefault(k, {"IP":0,"OP":0,"PH":0})
                for t in ("IP","OP","PH"): dd[t] += v[t]
        for k, v in c["op"].items():
            rec = op_agg.setdefault(k, {"new":0,"free":0,"renew":0,"tot":0})
            for fld in rec: rec[fld] += v[fld]
        for rec in c["dis"]:
            key = (rec["pid"], rec["date"])
            if key not in seen_pid:
                seen_pid.add(key); discharges.append(rec)
    if fcache_dirty:
        json.dump(fcache, open(fcache_path, "w"))

    # ---- Daily MIS files: doctor x month granularity ----
    mis_by_date = {}
    for f in glob.glob(os.path.join(folder, "**", "Daily MIS*.xlsx"), recursive=True):
        d = file_date(f)
        if d and (d not in mis_by_date or os.path.getmtime(f) > os.path.getmtime(mis_by_date[d])):
            mis_by_date[d] = f
    mis_month_files = {}
    for d in sorted(mis_by_date): mis_month_files[f"{d.year}-{d.month:02d}"] = (d, mis_by_date[d])

    mom_fy = {}
    mis_cache_path = os.path.join(tools, "mis_cache.json")
    mis_cache = {}
    if os.path.exists(mis_cache_path):
        try: mis_cache = json.load(open(mis_cache_path))
        except Exception: mis_cache = {}
    mis_budget = float(os.environ.get("MIS_BUDGET", "0")) or None
    _t_mis = time.time()
    mis_deferred = 0
    for mkey, (d, f) in sorted(mis_month_files.items(), reverse=True):
        sig = "%s|%d" % (os.path.basename(f), int(os.path.getmtime(f)))
        cached = mis_cache.get(mkey)
        if cached and cached.get("src") == sig and history.get(mkey, {}).get("doctors"):
            mom_fy.update(cached.get("mom") or {})
            continue
        if mis_budget and (time.time() - _t_mis) > mis_budget:
            mis_deferred += 1
            if cached: mom_fy.update(cached.get("mom") or {})
            continue
        print("Parsing Daily MIS", os.path.basename(f), flush=True)
        w = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sheets = {norm(s): s for s in w.sheetnames}
        def get(name): return w[sheets[norm(name)]] if norm(name) in sheets else None
        docs, doc_daily = {}, {}
        pairs = [("Doc wise revenue date conso.", "rev"), ("Patient visits", "opv"),
                 ("No. Admissions", "adm"), ("No. discharges", "dis")]
        for sn, field in pairs:
            ws = get(sn)
            if ws is None: continue
            totals, per_day = parse_doctor_day_matrix(ws, field)
            if field == "rev": doc_daily = per_day
            for (dept, doc), val in totals.items():
                rec = docs.setdefault(doc, {"dept": dept, "rev": 0, "opv": 0, "adm": 0, "dis": 0})
                if dept: rec["dept"] = dept
                rec[field] += val
        ws = get("MoM FY 26-27") or get("MoM FY 27-28")
        this_mom = parse_mom_fy(ws) if ws is not None else {}
        mom_fy.update(this_mom)
        w.close()
        mis_cache[mkey] = {"src": sig, "mom": this_mom}
        history[mkey] = {"asOf": d.strftime("%Y-%m-%d"),
                         "daysElapsed": d.day,
                         "doctors": {k: v for k, v in docs.items() if any(v[f] for f in ("rev","opv","adm","dis"))},
                         "docDaily": {k: {str(day): round(v, 0) for day, v in dd.items()}
                                      for k, dd in doc_daily.items() if dd}}
    json.dump(history, open(hist_path, "w"))
    json.dump(mis_cache, open(mis_cache_path, "w"))
    print("History months:", sorted(history), "| MIS deferred:", mis_deferred, flush=True)

    # discharge aggregates per doctor + overall
    dis_by_doc, status_mix, payer_mix = {}, {}, {}
    for rec in discharges:
        status_mix[rec["status"]] = status_mix.get(rec["status"], 0) + 1
        payer_mix[rec["scheme"]] = payer_mix.get(rec["scheme"], 0) + 1
        dd = dis_by_doc.setdefault(rec["doctor"], {"n": 0, "losSum": 0, "losN": 0, "status": {}, "cash": 0})
        dd["n"] += 1
        if rec["alos"] is not None: dd["losSum"] += rec["alos"]; dd["losN"] += 1
        dd["status"][rec["status"]] = dd["status"].get(rec["status"], 0) + 1
        if rec["scheme"] == "Cash": dd["cash"] += 1

    # ---- FY27 AOP plan (cached so the tab survives if the file moves) ----
    aop_cache = os.path.join(tools, "aop_fy27.json")
    aop = None
    ap = find_aop(folder)
    if ap:
        try:
            aop = parse_aop(ap)
            if aop: json.dump(aop, open(aop_cache, "w"))
            print("AOP parsed from", os.path.basename(ap),
                  "| FY27 plan IP+OP %.1f Cr" % (aop["fyPlanRev"] / 1e7))
        except Exception as e:
            print("AOP parse failed:", e)
    if not aop and os.path.exists(aop_cache):
        try:
            aop = json.load(open(aop_cache)); print("AOP from cache")
        except Exception: aop = None
    if not aop: print("WARNING: no FY27 AOP available; projection tab will hide plan columns")

    # ---- FY27 monthly financials (BRM deck actuals + MIS pack P&L, both units) ----
    fy27 = None
    if os.environ.get("SKIP_FY27") != "1":
        try:
            sys.path.insert(0, tools)
            import parse_fy27
            fy27 = parse_fy27.collect(folder, tools)
            if fy27:
                print("FY27: %d unit P&L, %d monthly blocks" %
                      (len(fy27.get("units", {})), len(fy27.get("monthly", {}))))
        except Exception as e:
            print("FY27 section skipped:", e)

    data = dict(
        generated=datetime.datetime.now().strftime("%d %b %Y %H:%M"),
        latestDate=latest_date.strftime("%d %b %Y"),
        filesParsed=[os.path.basename(by_date[d]) for d in dates] +
                    [os.path.basename(f) for _, f in mis_month_files.values()],
        yearTables=year_tables, daily=daily, summary=summary,
        deptTop=[{"name": k, "rev": v} for k, v in sorted(dept_tot.items(), key=lambda x: -x[1])[:14]],
        deptDates=dept_dates,
        history=history, momFY=mom_fy,
        disByDoc=dis_by_doc, statusMix=status_mix, payerMix=payer_mix,
        docTypeMix=doc_type_mix, opAgg=op_agg,
        disDates=sorted({r["date"] for r in discharges}),
        nDischarges=len(discharges),
        aop=aop,
        fy27=fy27,
        holidays=HOLIDAYS,
    )
    out = os.path.join(folder, "LHRC_Revenue_Dashboard.html")
    open(out, "w", encoding="utf-8").write(TEMPLATE.replace("__DATA__", json.dumps(data)))
    print("Wrote", out)

TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>VPS Lakeshore — Daily Revenue Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root{--blue:#2B7CBE;--maroon:#8B1A4A;--gray:#7F8C9B;--bg:#f6f8fa;--card:#fff;--good:#1a7f4e;--bad:#c0392b}
*{box-sizing:border-box}body{margin:0;font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;background:var(--bg);color:#243342}
header{background:linear-gradient(90deg,var(--blue),#1b5e94);color:#fff;padding:18px 28px;display:flex;justify-content:space-between;align-items:baseline;flex-wrap:wrap}
header h1{margin:0;font-size:21px;font-weight:600}
header .sub{font-size:12.5px;opacity:.85}
.wrap{max-width:1320px;margin:0 auto;padding:20px 24px 48px}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(165px,1fr));gap:12px;margin:6px 0 22px}
.card{background:var(--card);border-radius:10px;padding:13px 15px;box-shadow:0 1px 3px rgba(36,51,66,.08);border-top:3px solid var(--blue)}
.card.m{border-top-color:var(--maroon)}
.card .lbl{font-size:11px;color:var(--gray);text-transform:uppercase;letter-spacing:.05em}
.card .val{font-size:21px;font-weight:650;margin-top:3px}
.card .delta{font-size:12px;margin-top:2px}
.up{color:var(--good)}.dn{color:var(--bad)}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:18px}
@media(max-width:950px){.grid{grid-template-columns:1fr}}
.panel{background:var(--card);border-radius:10px;padding:16px 18px;box-shadow:0 1px 3px rgba(36,51,66,.08);margin-bottom:18px}
.panel h2{margin:0 0 4px;font-size:14.5px;color:var(--maroon);font-weight:650}
.panel .note{font-size:11.5px;color:var(--gray);margin-bottom:10px}
.panel canvas{max-height:330px}
table{width:100%;border-collapse:collapse;font-size:12px}
th{text-align:left;color:var(--gray);font-weight:600;border-bottom:2px solid #e3e8ee;padding:6px 7px;cursor:pointer;white-space:nowrap;user-select:none}
th:hover{color:var(--blue)}
td{padding:5px 7px;border-bottom:1px solid #eef1f5;white-space:nowrap}
td.r,th.r{text-align:right;font-variant-numeric:tabular-nums}
td.doc{max-width:190px;overflow:hidden;text-overflow:ellipsis}
.pill{display:inline-block;padding:2px 9px;border-radius:20px;font-size:11px;background:rgba(43,124,190,.1);color:var(--blue);margin:0 6px 6px 0}
.mbtn{border:1px solid #d4dbe3;background:#fff;color:#243342;border-radius:6px;padding:4px 12px;margin-right:6px;cursor:pointer;font-size:12.5px}
.mbtn.on{background:var(--blue);color:#fff;border-color:var(--blue)}
.tag{font-size:10.5px;padding:1px 7px;border-radius:10px}
.tag.g{background:rgba(26,127,78,.12);color:var(--good)}.tag.r{background:rgba(192,57,43,.1);color:var(--bad)}.tag.y{background:rgba(127,140,155,.15);color:#5a6875}
footer{font-size:11px;color:var(--gray);margin-top:26px;line-height:1.6}
.findgrid{display:grid;grid-template-columns:repeat(auto-fill,minmax(290px,1fr));gap:12px}
.fcard{border:1px solid #e3e8ee;border-left:4px solid var(--gray);border-radius:8px;padding:11px 13px;background:#fff;box-shadow:0 1px 2px rgba(36,51,66,.05)}
.fcard.g{border-left-color:var(--good)}.fcard.r{border-left-color:var(--bad)}.fcard.y{border-left-color:#c8952b}
.fcard .who{font-weight:650;font-size:12.5px}
.fcard .who .dp{font-weight:500;font-size:10.5px;color:var(--gray);margin-left:6px;text-transform:uppercase;letter-spacing:.03em}
.fcard .claim{font-size:12px;line-height:1.5;margin:5px 0 4px;color:#33475b}
.fcard .imp{font-size:14.5px;font-weight:700}
.fcard.g .imp{color:var(--good)}.fcard.r .imp{color:var(--bad)}.fcard.y .imp{color:#a37a1e}
.fcard .evid{font-size:10.5px;color:var(--gray);margin-top:4px;line-height:1.45}
.scroll{overflow-x:auto}
.tabs{background:#1b5e94;padding:0 28px;display:flex;gap:2px}
.tabs button{background:none;border:0;border-bottom:3px solid transparent;color:rgba(255,255,255,.72);
 padding:10px 18px;font-size:13px;font-weight:600;cursor:pointer;font-family:inherit}
.tabs button:hover{color:#fff}
.tabs button.on{color:#fff;border-bottom-color:#fff}
.view{display:none}.view.on{display:block}
.ctrls{display:flex;flex-wrap:wrap;gap:20px;align-items:center;background:#f0f5fa;
 border:1px solid #dde6ef;border-radius:8px;padding:12px 16px;margin-bottom:16px}
.ctrl{display:flex;flex-direction:column;gap:4px}
.ctrl .cl{font-size:10.5px;color:var(--gray);text-transform:uppercase;letter-spacing:.05em;font-weight:600}
.ctrl .cv{font-size:12.5px;font-weight:650;color:var(--blue)}
.ctrl input[type=range]{width:170px;accent-color:var(--blue)}
.sw{display:flex;align-items:center;gap:7px;font-size:12.5px;cursor:pointer;user-select:none}
.rbtn{border:1px solid #d4dbe3;background:#fff;color:#243342;border-radius:6px;padding:3px 10px;
 cursor:pointer;font-size:12px;font-family:inherit}
.rbtn.on{background:var(--blue);color:#fff;border-color:var(--blue)}
tr.pastm td{background:#fafcfe;color:#5a6875}
.vzn{border-top:3px solid var(--maroon)}
.vzn h3{margin:0 0 6px;font-size:12.5px;color:var(--blue);font-weight:700;letter-spacing:.02em}
.vzn p{margin:0 0 9px;font-size:12.5px;line-height:1.62;color:#243342}
.vzn p.vfoot{font-size:11px;color:var(--gray);margin:8px 0 0}
.vzn p.vwarn{font-size:11.5px;color:#5a6875;background:rgba(127,140,155,.1);border-left:3px solid var(--gray);padding:8px 11px;border-radius:0 6px 6px 0;margin-top:10px}
.vstrip{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:12px 0 16px}
@media(max-width:950px){.vstrip{grid-template-columns:1fr 1fr}}
.vstat{background:rgba(43,124,190,.06);border-radius:8px;padding:10px 12px}
.vlbl{font-size:10.5px;color:var(--gray);text-transform:uppercase;letter-spacing:.04em}
.vval{font-size:22px;font-weight:700;line-height:1.25;font-variant-numeric:tabular-nums}
.vval.bad{color:var(--bad)}.vval.good{color:var(--good)}
.vsub{font-size:11px;color:var(--gray)}
.vgrid{display:grid;grid-template-columns:1fr 1fr;gap:18px}
@media(max-width:950px){.vgrid{grid-template-columns:1fr}}
.vbox{background:#fbfcfd;border:1px solid #e8edf2;border-radius:8px;padding:13px 15px}
.vbox.read{border-left:3px solid var(--maroon)}
table.vtab{margin-top:4px}
table.vtab th{cursor:default;padding:5px 6px}
table.vtab th:hover{color:var(--gray)}
table.vtab td{padding:4px 6px}
table.vtab td.good{color:var(--good);font-weight:650}
table.vtab td.bad{color:var(--bad);font-weight:650}
table.vtab td.flat{color:var(--gray)}
table.vtab tr.sep td{border-top:2px solid #dfe6ed}
tr.fytot td{font-weight:700;border-top:2px solid #d4dbe3;background:#f4f8fc}
</style></head><body>
<header><div><h1>VPS Lakeshore · Daily Revenue Dashboard</h1>
<div class="sub">Lakeshore Hospital &amp; Research Centre Ltd, Kochi · Global Lifecare</div></div>
<div class="sub" id="asof"></div></header>
<nav class="tabs">
<button id="tabOps" class="on" onclick="showView('ops')">Daily operations</button>
<button id="tabProj" onclick="showView('proj')">FY 26-27 projection</button>
<button id="tabCmi" onclick="showView('cmi')">Case mix (CMI)</button>
<button id="tabFy27" onclick="showView('fy27')">FY27 financials</button>
</nav>
<div class="wrap view on" id="viewOps">
<div class="cards" id="cards"></div>

<div id="mtdStrip"></div>
<div class="panel vzn" id="vznPanel">
<h2 id="vznTitle">Month-on-month variance</h2>
<div class="note" id="vznNote"></div>

<div class="vstrip" id="vznStrip"></div>
<div class="vgrid" id="vznGrid"></div>
</div>

<div class="panel"><h2>Daily Gross Revenue — Actual vs Budget</h2>
<div class="note">OP / IP / Pharmacy stacked; line = budgeted total.</div>
<div id="mbtns" style="margin-bottom:8px"></div><canvas id="dailyChart"></canvas></div>

<div class="grid">
<div class="panel"><h2>Monthly Revenue — FY 26-27 vs Budget vs FY 25-26</h2>
<div class="note">₹ Cr. Current-month bar is month-to-date.</div><canvas id="monthlyChart"></canvas></div>
<div class="panel"><h2>Efficiency — ARPOB · Occupancy · ALOS</h2>
<div class="note">ARPOB = gross revenue ÷ occupied bed-days (₹ ’000/day). Bars = occupancy %, dotted = ALOS (days).</div><canvas id="effChart"></canvas></div>
<div class="panel"><h2>Collections — Cash + Credit vs Budget</h2>
<div class="note">₹ Cr per month, FY 26-27. Line marker = collections ÷ gross revenue (cash conversion).</div><canvas id="collChart"></canvas></div>
<div class="panel"><h2>Volumes — OP Visits &amp; IP Discharges</h2>
<div class="note">FY 26-27 monthly vs budget (dotted).</div><canvas id="volChart"></canvas></div>
</div>

<div class="panel"><h2>Doctor Findings — Auto-Detected Signals</h2>
<div class="note" id="findnote"></div>
<div class="findgrid" id="findings"></div></div>

<div class="panel"><h2>Doctor League Table — Month on Month</h2>
<div class="note" id="dlnote"></div>
<div style="margin-bottom:8px"><input id="docFilter" placeholder="Filter doctor / department…" style="padding:5px 10px;border:1px solid #d4dbe3;border-radius:6px;width:260px;font-size:12.5px"></div>
<div class="scroll"><table id="league"><thead><tr>
<th data-k="doc">Doctor</th><th data-k="dept">Department</th>
<th class="r" data-k="revC">Rev cur (₹L)</th><th class="r" data-k="rrC">₹L/day cur</th>
<th class="r" data-k="revP">Rev prev (₹L)</th><th class="r" data-k="rrP">₹L/day prev</th>
<th class="r" data-k="mom">Δ run-rate</th>
<th class="r" data-k="share">Share</th><th class="r" data-k="rkD">Rank Δ</th>
<th class="r" data-k="opC">OP cur</th><th class="r" data-k="opP">OP prev</th>
<th class="r" data-k="admC">Adm cur</th><th class="r" data-k="disC">Disch cur</th>
<th class="r" data-k="conv">Conv %</th><th class="r" data-k="ipmix">IP mix*</th>
<th class="r" data-k="newp">New OP*</th><th class="r" data-k="freep">Free OP*</th>
<th class="r" data-k="rpd">₹L/disch</th><th class="r" data-k="alos">ALOS*</th>
<th class="r" data-k="recov">Recovered*</th><th class="r" data-k="dama">DAMA*</th><th class="r" data-k="exp">Expired*</th>
<th class="r" data-k="cash">Cash mix*</th>
</tr></thead><tbody></tbody></table></div></div>

<div class="panel"><h2>Doctor × Day Revenue Matrix</h2>
<div class="note" id="mxnote"></div>
<div style="margin-bottom:8px">
<span id="mxbtns"></span>
<input id="mxFilter" placeholder="Filter doctor…" style="padding:5px 10px;border:1px solid #d4dbe3;border-radius:6px;width:220px;font-size:12.5px;margin-left:8px">
</div>
<div class="scroll" style="max-height:520px;overflow:auto"><table id="matrix"><thead></thead><tbody></tbody></table></div></div>

<div class="panel"><h2>Department League Table — Month on Month</h2>
<div class="note" id="dtnote"></div>
<div class="scroll"><table id="deptTable"><thead><tr>
<th data-k="dept">Department</th>
<th class="r" data-k="revC">Rev cur (₹L)</th><th class="r" data-k="rrC">₹L/day cur</th>
<th class="r" data-k="revP">Rev prev (₹L)</th><th class="r" data-k="rrP">₹L/day prev</th>
<th class="r" data-k="revLY">LY same mo (₹L)</th><th class="r" data-k="revLYfy">LY FY total (₹L)</th>
<th class="r" data-k="mom">Δ run-rate</th><th class="r" data-k="share">Share</th>
<th class="r" data-k="opC">OP cur</th><th class="r" data-k="admC">Adm cur</th>
<th class="r" data-k="disC">Disch cur</th><th class="r" data-k="conv">Conv %</th>
<th class="r" data-k="rpd">₹L/disch</th><th class="r" data-k="alos">ALOS*</th>
<th class="r" data-k="arpob">ARPOB† ₹k</th><th class="r" data-k="cash">Cash mix*</th>
</tr></thead><tbody></tbody></table></div></div>

<div class="grid">
<div class="panel"><h2>Doctor Concentration — Cumulative Revenue Share</h2>
<div class="note" id="concnote"></div><canvas id="concChart"></canvas></div>
<div class="panel"><h2>Department Revenue — Current vs Previous Month (run-rate)</h2>
<div class="note">₹ Lakhs per day, doctor-attributed revenue from Daily MIS.</div><canvas id="deptMoM" style="max-height:400px"></canvas></div>
<div class="panel"><h2>Discharge Outcomes &amp; Payer Mix</h2>
<div class="note" id="disnote"></div>
<div style="display:flex;gap:10px"><div style="width:50%"><canvas id="statusChart"></canvas></div>
<div style="width:50%"><canvas id="payerChart"></canvas></div></div></div>
<div class="panel"><h2>Department Revenue — Captured Flash Days</h2>
<div class="note" id="deptnote"></div><canvas id="deptChart" style="max-height:400px"></canvas></div>
</div>
</div>

<div class="wrap view" id="viewProj">
<div class="panel" style="border-top:3px solid var(--maroon)">
<h2>Projection assumptions</h2>
<div class="note" id="projBasisNote">Closed months are actual. The current month is month-to-date banked revenue plus remaining days at the run-rate. Later months are run-rate &times; calendar days, flexed by the options below.</div>
<div class="ctrls">
 <div class="ctrl"><span class="cl">Run-rate window</span>
  <span id="rrBtns"></span></div>
 <div class="ctrl"><span class="cl">Run-rate used</span><span class="cv" id="rrVal"></span></div>
 <div class="ctrl"><span class="cl">Monthly ramp</span>
  <input type="range" id="rampS" min="-2" max="5" step="0.25" value="1.25">
  <span class="cv" id="rampVal"></span></div>
 <div class="ctrl"><span class="cl">Seasonality</span>
  <label class="sw"><input type="checkbox" id="seasS" checked> Apply FY 25-26 monthly shape</label></div>
</div>
<div class="cards" id="pcards" style="margin:0"></div>
</div>

<div class="panel"><h2>FY 26-27 Revenue — Actual, Projected and AOP Plan</h2>
<div class="note" id="pchartnote"></div><canvas id="projChart" style="max-height:360px"></canvas></div>

<div class="grid">
<div class="panel"><h2>Cumulative Landing Zone vs Plan</h2>
<div class="note">Cumulative ₹ Cr through the year. Shaded gap = projection minus AOP plan.</div>
<canvas id="cumChart"></canvas></div>
<div class="panel"><h2>Scenario Fan — FY Total by Ramp</h2>
<div class="note">FY 26-27 total at each monthly ramp setting, holding the selected run-rate window and seasonality. Marker = current setting.</div>
<canvas id="fanChart"></canvas></div>
</div>

<div class="panel"><h2>Monthly Build</h2>
<div class="note" id="ptabnote"></div>
<div class="scroll"><table id="projTable"><thead><tr>
<th>Month</th><th class="r">Days</th><th class="r">Seas. idx</th>
<th class="r">FY 25-26 (₹ Cr)</th><th class="r">AOP plan (₹ Cr)</th>
<th class="r">Actual / projected (₹ Cr)</th><th class="r">₹ Cr/day</th>
<th class="r">vs AOP (₹ Cr)</th><th class="r">vs FY 25-26 (₹ Cr)</th><th>Basis</th>
</tr></thead><tbody></tbody></table></div></div>

<div class="panel"><h2>Read-across</h2>
<div class="note">Auto-generated from the current assumption set.</div>
<div id="pcomment" style="font-size:12.5px;line-height:1.65;color:#33475b"></div></div>
</div>

<div class="wrap view" id="viewCmi">
<div class="panel" style="border-top:3px solid var(--maroon)">
<h2>How this CMI is built</h2>
<div class="note" id="cmiMethod"></div>
<div class="ctrls" style="margin-top:10px">
 <div class="ctrl"><span class="cl">Base for the hospital average</span>
  <label class="sw"><input type="checkbox" id="cmiNoDay"> Exclude day-care specialties (ALOS &lt; 2 d)</label></div>
 <div class="ctrl"><span class="cl">Cases in base</span><span class="cv" id="cmiBaseN"></span></div>
</div>
<div class="cards" id="cmiCards" style="margin:0"></div>
</div>

<div class="grid">
<div class="panel"><h2>Case Mix Index by Specialty</h2>
<div class="note">CMI = specialty ₹ per inpatient case ÷ hospital average. 1.00 = average acuity. Bars above 1.0 in blue, below in maroon.</div>
<canvas id="cmiChart" style="max-height:520px"></canvas></div>
<div class="panel"><h2>Case-Mix Map — Acuity vs Volume</h2>
<div class="note">X = CMI, Y = discharges, bubble area = IP-attributed revenue. Top-right = high acuity at scale; bottom-right = high acuity, thin volume.</div>
<canvas id="cmiScatter" style="max-height:520px"></canvas></div>
</div>

<div class="panel"><h2>Specialty Case Mix — Detail</h2>
<div class="note" id="cmiTabNote"></div>
<div class="scroll"><table id="cmiTable"><thead><tr>
<th data-k="dept">Specialty</th>
<th class="r" data-k="iprev">IP rev (₹ Cr)</th><th class="r" data-k="dis">Discharges</th>
<th class="r" data-k="alos">ALOS</th><th class="r" data-k="rpc">₹L / case</th>
<th class="r" data-k="cmi">CMI</th><th class="r" data-k="rpb">₹L / bed-day</th>
<th class="r" data-k="ii">Intensity idx</th>
<th class="r" data-k="shRev">Rev share</th><th class="r" data-k="shDis">Case share</th>
<th class="r" data-k="mom">Δ ₹L/case MoM</th><th class="r" data-k="docs">Doctors</th>
</tr></thead><tbody></tbody></table></div></div>

<div class="panel"><h2>Doctor-Level Case Mix</h2>
<div class="note">Same construction at doctor level. Doctors with fewer than 5 discharges in the month are hidden — per-case figures are too noisy to read below that.</div>
<div style="margin-bottom:8px">
<input id="cmiDocFilter" placeholder="Filter doctor / specialty…" style="padding:5px 10px;border:1px solid #d4dbe3;border-radius:6px;width:260px;font-size:12.5px">
<span id="cmiMinBtns" style="margin-left:8px"></span></div>
<div class="scroll" style="max-height:560px;overflow:auto"><table id="cmiDocTable"><thead><tr>
<th data-k="doc">Doctor</th><th data-k="dept">Specialty</th>
<th class="r" data-k="iprev">IP rev (₹ L)</th><th class="r" data-k="dis">Disch</th>
<th class="r" data-k="alos">ALOS</th><th class="r" data-k="rpc">₹L / case</th>
<th class="r" data-k="cmi">CMI</th><th class="r" data-k="rpb">₹L / bed-day</th>
<th class="r" data-k="ii">Intensity idx</th><th class="r" data-k="cash">Cash mix</th>
</tr></thead><tbody></tbody></table></div></div>

<div class="panel"><h2>Excluded from the index</h2>
<div class="note">Service and non-admitting departments carry attributed revenue but few or no discharges of their own, so a per-case figure would be meaningless. They are excluded from the index and from the hospital average, but their revenue is listed here so the reconciliation is visible.</div>
<div id="cmiExcl" style="font-size:12px"></div></div>

<div class="panel"><h2>Read-across</h2>
<div class="note">Auto-generated from the current month's case mix.</div>
<div id="cmiComment" style="font-size:12.5px;line-height:1.65;color:#33475b"></div></div>
</div>

<div class="wrap view" id="viewFy27">
<div class="panel vzn">
<h2>FY 26-27 monthly financials &mdash; <span id="f27Through"></span></h2>
<div class="note" id="f27Basis"></div>
<div id="f27UnitBtns" style="margin:10px 0 4px"></div>
</div>
<div class="cards" id="f27Cards"></div>
<div class="panel"><h2>Profit &amp; loss to EBITDA</h2>
<div class="note" id="f27PlNote"></div>
<div class="scroll"><table class="vtab" id="f27Pl"><thead></thead><tbody></tbody></table></div></div>
<div class="grid">
<div class="panel"><h2>Monthly revenue &mdash; actual</h2>
<div class="note">&#8377; Cr. OP / IP split, actuals only (see basis note above).</div>
<canvas id="f27RevChart"></canvas></div>
<div class="panel"><h2>OP visits &mdash; new vs revisit</h2>
<div class="note">Monthly actuals.</div>
<canvas id="f27OpvChart"></canvas></div>
</div>
<div class="grid">
<div class="panel"><h2>IP discharges &amp; occupancy</h2>
<div class="note">Group level (Kochi + Calicut). Bars = discharges, line = occupancy %.</div>
<canvas id="f27OccChart"></canvas></div>
<div class="panel"><h2>Payor mix</h2>
<div class="note">&#8377; Cr, group level, monthly actuals.</div>
<canvas id="f27PayChart"></canvas></div>
</div>
<div id="f27TrackerWrap">
<div class="panel vzn"><h2>AOP tracker &mdash; <span id="trkSrc"></span></h2>
<div class="note" id="trkNote"></div>
<div class="vstrip" id="trkStrip"></div>
</div>

<div class="panel"><h2>EBITDA &mdash; plan vs actual</h2>
<div class="note">&#8377; Cr. Bars = actual, line = AOP plan. Marker labels show EBITDA margin on actual revenue.</div>
<canvas id="trkEbChart"></canvas></div>

<div class="panel vzn"><h2>Two <span id="trkAugMon">monthly</span> plan figures</h2>
<div class="note" id="trkAugNote"></div></div>

<div class="panel"><h2>P&amp;L detail &mdash; plan vs actual</h2>
<div class="note" id="trkPlNote"></div>
<div style="margin-bottom:8px">
 <button class="rbtn on" id="trkSortVar" onclick="trkSort('var')">Sort by variance</button>
 <button class="rbtn" id="trkSortOrd" onclick="trkSort('ord')">P&amp;L order</button>
</div>
<div class="scroll"><table class="vtab" id="trkPl"><thead></thead><tbody></tbody></table></div></div>

<div class="panel"><h2>Initiative build-up</h2>
<div class="note" id="trkInitNote"></div>
<div class="scroll" style="max-height:520px;overflow:auto"><table class="vtab" id="trkInit"><thead></thead><tbody></tbody></table></div></div>
</div>

<div class="panel vzn"><h2>Not available in the FY27 packs</h2>
<div class="note" id="f27Gaps"></div></div>
</div>

<div class="wrap" style="padding-top:0"><footer id="foot"></footer></div>
<script>
const D = __DATA__;
const CR=1e7, L=1e5;
const fmtCr=v=>'₹'+(v/CR).toFixed(2)+' Cr';
const pct=(a,b)=>b? ((a/b-1)*100):0;
const BLUE='#2B7CBE',MAROON='#8B1A4A',GRAY='#7F8C9B',LT='rgba(43,124,190,.35)',GOLD='#c8952b';
const tc=s=>s.split(' ').map(w=>w.includes('.')&&w.length<=5?w:w.charAt(0)+w.slice(1).toLowerCase()).join(' ');
document.getElementById('asof').textContent='Data through '+D.latestDate+' · generated '+D.generated;

// value labels on bars: stacked bars show one overall total; plain/grouped bars show their value; Budget bars skipped
const valLabel={id:'valLabel',afterDatasetsDraw(chart){
 const metas=chart.getSortedVisibleDatasetMetas().filter(m=>m.type==='bar'&&!/budget/i.test(chart.data.datasets[m.index].label||''));
 if(!metas.length)return;
 const ctx=chart.ctx;ctx.save();
 const horiz=chart.options.indexAxis==='y';
 const n=(chart.data.labels||[]).length;
 const rot=!horiz&&n>14;
 ctx.font='600 '+(n>14?'8.5':'10')+'px -apple-system,Segoe UI,Roboto,Arial';
 ctx.fillStyle='#3d4a57';
 const fv=v=>{const a=Math.abs(v);return a>=1000?Math.round(v).toLocaleString('en-IN'):a>=100?v.toFixed(0):v.toFixed(1)};
 const draw=(val,el)=>{
  if(!val||!el)return;
  if(horiz){ctx.textAlign='left';ctx.textBaseline='middle';ctx.fillText(fv(val),el.x+4,el.y);}
  else if(rot){ctx.translate(el.x,el.y-3);ctx.rotate(-Math.PI/2);ctx.textAlign='left';ctx.textBaseline='middle';ctx.fillText(fv(val),0,0);ctx.setTransform(1,0,0,1,0,0);}
  else{ctx.textAlign='center';ctx.textBaseline='bottom';ctx.fillText(fv(val),el.x,el.y-3);}
 };
 const stacks={},plain=[];
 metas.forEach(m=>{const s=chart.data.datasets[m.index].stack;if(s)(stacks[s]=stacks[s]||[]).push(m);else plain.push(m);});
 Object.values(stacks).forEach(ms=>{
  const top=ms[ms.length-1];
  top.data.forEach((el,i)=>{let t=0;ms.forEach(m=>{t+=+chart.data.datasets[m.index].data[i]||0});draw(t,el);});
 });
 plain.forEach(m=>{m.data.forEach((el,i)=>draw(+chart.data.datasets[m.index].data[i],el));});
 ctx.restore();
}};
Chart.register(valLabel);

const fyKeys=Object.keys(D.yearTables);
const cur=D.yearTables[fyKeys[0]]||{months:[]}, prev=D.yearTables[fyKeys[1]]||{months:[]};
const mwd=cur.months.filter(m=>m.revTot>0);
const ytd={rev:0,bud:0,coll:0,budColl:0,opv:0,ipd:0};
mwd.forEach(m=>{ytd.rev+=m.revTot;ytd.bud+=m.budTot;ytd.coll+=m.collTot;ytd.budColl+=m.budCollTot;ytd.opv+=m.opVisits;ytd.ipd+=m.ipDisch;});
const mtd=mwd[mwd.length-1]||{};
const prevYtdRev=prev.months.slice(0,mwd.length).reduce((a,m)=>a+m.revTot,0);

// history months
const hMonths=Object.keys(D.history).sort();
const curM=hMonths[hMonths.length-1], prevM=hMonths[hMonths.length-2];
const hCur=curM? D.history[curM]:null, hPrev=prevM? D.history[prevM]:null;
function daysIn(mk,h){const [y,m]=mk.split('-').map(Number);
 const last=new Date(y,m,0).getDate();
 const asOfDay=h&&h.daysElapsed? h.daysElapsed:last;
 return Math.min(asOfDay,last);}
const dCur=curM? daysIn(curM,hCur):1, dPrev=prevM? daysIn(prevM,hPrev):1;
const mName=mk=>{const [y,m]=mk.split('-');return new Date(y,m-1,1).toLocaleString('en',{month:'short'})+'-'+y.slice(2)};

// efficiency series (FY26-27 from MoM sheet + fallback to flash occDays for prior FY)
const effM=Object.keys(D.momFY).sort();
const eff=effM.map(mk=>{const e=D.momFY[mk];
 const ym=mk; let rev=0;
 const mi=cur.months[(parseInt(mk.split('-')[1])-4+12)%12];
 if(mi) rev=mi.revTot;
 return {mk, arpob: e.occDays? rev/e.occDays/1000:null,
         occ: e.bedCap? e.occDays/e.bedCap*100:null, alos:e.alos||null};});

// KPI cards
function card(lbl,val,delta,cls,m){return `<div class="card${m?' m':''}"><div class="lbl">${lbl}</div><div class="val">${val}</div><div class="delta ${cls}">${delta}</div></div>`}
const dv=(a,b,sfx)=>{const p=pct(a,b);return [(p>=0?'▲ +':'▼ ')+p.toFixed(1)+'% '+sfx,p>=0?'up':'dn']};
let [d1,c1]=dv(mtd.revTot,mtd.budTot,'vs budget');
let [d2,c2]=dv(ytd.rev,ytd.bud,'vs budget');
let [d3,c3]=dv(ytd.rev,prevYtdRev,'YoY');
let [d4,c4]=dv(ytd.coll,ytd.budColl,'vs budget');
const bo=(D.summary.bedOcc||[0,0,0]);
const lastEff=eff.filter(e=>e.arpob).slice(-1)[0]||{};
const prevEff=eff.filter(e=>e.arpob).slice(-2)[0]||{};
const conv=(D.summary.real&&D.summary.real['IP Conv. Rate'])||[0,0];
// month run-rate projection: MTD ÷ elapsed working days × total working days, vs full-month budget
const wdTot=D.summary.workingDays||0, wdCum=D.summary.cumDays||0;
const dailyKeys0=Object.keys(D.daily).sort();
const curKey=dailyKeys0[dailyKeys0.length-1];
const fullBud=D.summary.monthBudget||(curKey? D.daily[curKey].reduce((a,r)=>a+r.budTot,0):0);
const rrDay=wdCum? (mtd.revTot||0)/wdCum:0;
const proj=rrDay*wdTot;
let projCard='';
if(proj&&fullBud){
 const p=pct(proj,fullBud);
 projCard=card('Month Run-Rate → Landing',fmtCr(proj),
  `₹${(rrDay/CR).toFixed(2)} Cr/wk-day × ${wdTot} days · <span class="${p>=0?'up':'dn'}">${p>=0?'▲ +':'▼ '}${p.toFixed(1)}% vs ₹${(fullBud/CR).toFixed(0)} Cr budget</span>`,'',1);
}
document.getElementById('cards').innerHTML=
 projCard+
 card('MTD Revenue ('+(mtd.month||'')+')',fmtCr(mtd.revTot||0),d1,c1)+
 card('YTD Revenue FY 26-27',fmtCr(ytd.rev),d2,c2)+
 card('YTD YoY',fmtCr(prevYtdRev)+' LY',d3,c3,1)+
 card('YTD Collections',fmtCr(ytd.coll),d4,c4,1)+
 card('ARPOB (MTD)','₹'+(lastEff.arpob||0).toFixed(1)+'k','prev mo ₹'+(prevEff.arpob||0).toFixed(1)+'k','')+
 card('ALOS (MTD)',(lastEff.alos||0).toFixed(1)+' d','prev mo '+(prevEff.alos||0).toFixed(1)+' d','')+
 card('Bed Occupancy (MTD)',(bo[1]*100).toFixed(0)+'%','YTD '+(bo[2]*100).toFixed(0)+'%','')+
 card('OP→IP Conversion',(conv[0]*100).toFixed(1)+'%','YTD '+(conv[1]*100).toFixed(1)+'%','');

// ===========================================================================
// Auto-generated month-on-month variance panel.
// Compares the latest CLOSED month against the previous month with data.
// Every figure, table row and sentence below is derived from D at render
// time — nothing here is hand-written, so it cannot go stale.
// ===========================================================================
(function(){
 const HOL=D.holidays||{};
 const ACR=/^(LHRC|GI|OP|IP|ICU|CT|MRI|OB|ENT|NICU|PICU|CCU)$/;
 const nmT=s=>String(s).split(' ').map(w=>ACR.test(w)?w:tc(w)).join(' ');
 const MONF=['January','February','March','April','May','June','July',
             'August','September','October','November','December'];
 const dim=mk=>{const [y,m]=mk.split('-').map(Number);return new Date(y,m,0).getDate()};
 const mLong=mk=>{const [y,m]=mk.split('-').map(Number);return MONF[m-1]};
 const sgn=v=>(v>=0?'+':'−')+Math.abs(v).toFixed(1);
 const cls=(v,inv)=>{const t=inv?-v:v;return t>1.5?'good':t<-1.5?'bad':'flat'};
 const fL=v=>'₹'+v.toFixed(2)+' L';
 const fLd=v=>'₹'+v.toFixed(1)+' L';
 const iN=v=>Math.round(v).toLocaleString('en-IN');

 // ---- pick the months -----------------------------------------------------
 const keys=Object.keys(D.daily).filter(k=>(D.daily[k]||[]).length).sort();
 const closed=keys.filter(k=>D.daily[k].length>=dim(k));
 const tgt=closed[closed.length-1];
 if(!tgt){document.getElementById('vznPanel').style.display='none';return;}
 const base=keys.filter(k=>k<tgt).pop();
 const liveK=keys[keys.length-1];
 const live=(liveK>tgt)? liveK : null;

 // ---- per-month statistics ----------------------------------------------
 function stats(mk){
  if(!mk) return null;
  const R=D.daily[mk]||[], H=D.history[mk]||null, E=D.momFY[mk]||null;
  const n=R.length;
  const isSun=r=>r.dow==='Sun', isHol=r=>!!HOL[r.date]&&!isSun(r);
  const wk=R.filter(r=>!isSun(r)&&!isHol(r));
  const sum=(a,f)=>a.reduce((s,r)=>s+(+r[f]||0),0);
  const S={mk,n,R,H,E,
    label:mLong(mk),
    complete:n>=dim(mk),
    sundays:R.filter(isSun).length,
    hols:R.filter(isHol).map(r=>HOL[r.date]),
    holDates:R.filter(isHol).map(r=>r.date),
    wkDays:wk.length,
    rev:sum(R,'revTot'), bud:sum(R,'budTot'),
    revOP:sum(R,'revOP'), revIP:sum(R,'revIP'), revPH:sum(R,'revPH'),
    opv:sum(R,'opVisits'), ipd:sum(R,'ipDisch'),
    coll:sum(R,'collTot'),
    wkRev:sum(wk,'revTot')};
  S.wkOpv=sum(wk,'opVisits');
  S.perDay=S.rev/n; S.wkPerDay=S.wkDays? S.wkRev/S.wkDays:0;
  S.wkOpvPerDay=S.wkDays? S.wkOpv/S.wkDays:0;
  S.opPerDay=S.revOP/n; S.ipPerDay=S.revIP/n; S.phPerDay=S.revPH/n;
  S.opvPerDay=S.opv/n; S.ipdPerDay=S.ipd/n;
  S.revPerOpv=S.opv? S.revOP/S.opv:0;
  S.revPerDisch=S.ipd? S.revIP/S.ipd:0;
  S.occ=(E&&E.bedCap)? E.occDays/E.bedCap*100:null;
  S.arpob=(E&&E.occDays)? S.rev/E.occDays/1000:null;
  S.alos=(E&&E.alos)||null;
  S.conv=S.rev? S.coll/S.rev*100:null;
  // halves
  const h=Math.ceil(n/2);
  const f=R.slice(0,h), b=R.slice(h);
  S.h1=f.length? sum(f,'revTot')/f.length:0;
  S.h2=b.length? sum(b,'revTot')/b.length:0;
  S.h1n=f.length; S.h2n=b.length;
  // doctor / department aggregates
  S.docs={}; S.depts={}; S.adm=0;
  if(H&&H.doctors){
   Object.entries(H.doctors).forEach(([nm,v])=>{
    S.docs[nm]={rev:v.rev||0,dept:v.dept||'',opv:v.opv||0,adm:v.adm||0,dis:v.dis||0};
    S.adm+=(v.adm||0);
    const d=v.dept||'Unallocated';
    S.depts[d]=(S.depts[d]||0)+(v.rev||0);
   });
  }
  S.docRev=Object.values(S.docs).reduce((a,d)=>a+d.rev,0);
  S.admPerDay=S.adm/n;
  return S;
 }
 const A=stats(tgt), B=stats(base), Lv=stats(live);
 window.__closedMonth=A.label;   // consumed by the AOP-tracker plan-phasing note

 // ---- headline framing ---------------------------------------------------
 const dHead=B? pct(A.perDay,B.perDay):0;                 // per calendar day
 const dWork=(B&&B.wkPerDay)? pct(A.wkPerDay,B.wkPerDay):0; // ex-Sun/holiday
 const dBud=pct(A.rev,A.bud);
 const nwA=A.sundays+A.hols.length, nwB=B? B.sundays+B.hols.length:0;
 const up=dHead>=0;
 document.getElementById('vznTitle').textContent=
   B? ('Why '+A.label+' came in '+(up?'ahead of ':'behind ')+B.label)
     : (A.label+' — month in review');

 // how much of the headline gap the calendar explains
 let calShare=null;
 if(B&&Math.abs(dHead)>0.5&&Math.abs(dHead)>Math.abs(dWork))
   calShare=Math.min(100,Math.max(0,(1-dWork/dHead)*100));

 const holTxt=A.hols.length? (' plus '+A.hols.join(' and ')):'';
 const nwPhrase=nwA+' non-working day'+(nwA===1?'':'s')+' ('+A.sundays+
   ' Sunday'+(A.sundays===1?'':'s')+holTxt+')';
 let note='Generated from the data on every build — '+A.label+
   ' is closed at '+A.n+' of '+dim(tgt)+' days'+
   (B? ', compared against '+B.label+(B.complete?'':' ('+B.n+' days of data only)'):'')+'. ';
 if(B){
  note+='Calendar mix: '+A.label+' carries <b>'+nwPhrase+'</b> against '+nwB+' in '+B.label+
   (calShare!=null? ', which accounts for roughly '+calShare.toFixed(0)+'% of the headline gap':'')+
   '. The ex-Sunday, ex-holiday run-rate moved <b>'+sgn(dWork)+'%</b>. ';
  note+='Within '+A.label+', days 1–'+A.h1n+' ran at '+fLd(A.h1/L)+'/day against '+
   fLd(A.h2/L)+'/day over days '+(A.h1n+1)+'–'+A.n+
   ' ('+sgn(pct(A.h2,A.h1))+'%), where '+B.label+' went '+fLd(B.h1/L)+' → '+fLd(B.h2/L)+'.';
 }
 document.getElementById('vznNote').innerHTML=note;

 // ---- stat strip ---------------------------------------------------------
 const st=(l,v,s,c)=>'<div class="vstat"><div class="vlbl">'+l+'</div><div class="vval '+
   (c||'')+'">'+v+'</div><div class="vsub">'+s+'</div></div>';
 document.getElementById('vznStrip').innerHTML=
  (B? st('Run-rate vs full '+B.label,sgn(dHead)+'%',
      fLd(B.perDay/L)+'/day → '+fLd(A.perDay/L)+'/day ('+
      (A.perDay>=B.perDay?'+':'−')+fLd(Math.abs(A.perDay-B.perDay)/L)+'/day)',
      cls(dHead)):'')+
  (B? st('Ex-Sunday / holiday run-rate',sgn(dWork)+'%',
      fLd(B.wkPerDay/L)+' → '+fLd(A.wkPerDay/L)+'/day across '+A.wkDays+' working days',
      cls(dWork)):'')+
  st(A.label+' vs budget',sgn(dBud)+'%',
     fmtCr(A.rev)+' vs '+fmtCr(A.bud)+' ('+(A.rev>=A.bud?'+':'−')+
     fmtCr(Math.abs(A.rev-A.bud))+')',cls(dBud))+
  (A.conv!=null? st('Cash conversion',A.conv.toFixed(0)+'%',
     fmtCr(A.coll)+' collected on '+fmtCr(A.rev)+' gross'+
     (B&&B.conv!=null? ' · '+B.label+' '+B.conv.toFixed(0)+'%':''),
     (B&&B.conv!=null)?cls(A.conv-B.conv):''):'');

 // ---- box 1: rate vs volume ---------------------------------------------
 const rows=[
  ['OP revenue per visit', A.revPerOpv, B&&B.revPerOpv, v=>'₹'+iN(v), 0, 0],
  ['ARPOB (₹’000/bed-day)', A.arpob, B&&B.arpob, v=>v.toFixed(1), 0, 0],
  ['IP revenue per discharge', A.revPerDisch, B&&B.revPerDisch, v=>fL(v/L), 0, 0],
  ['ALOS (days)', A.alos, B&&B.alos, v=>v.toFixed(2), 0, 0],
  ['OP visits / day', A.opvPerDay, B&&B.opvPerDay, v=>iN(v), 1, 0],
  ['Bed occupancy', A.occ, B&&B.occ, v=>v.toFixed(1)+'%', 1, 1],
  ['IP discharges / day', A.ipdPerDay, B&&B.ipdPerDay, v=>v.toFixed(1), 1, 0],
  ['Admissions / day', A.admPerDay, B&&B.admPerDay, v=>v.toFixed(1), 1, 0]
 ].filter(r=>r[1]!=null&&isFinite(r[1]));
 const firstVol=rows.find(x=>x[4]===1);

 let tb='';
 rows.forEach(r=>{
  const [lbl,a,b,f,isVol,pts]=r;
  const has=(b!=null&&isFinite(b)&&b!==0);
  const dp=has? pct(a,b):null;
  const dCell=!has? '—'
    : pts? ((a>=b?'+':'−')+Math.abs(a-b).toFixed(1)+' pts')
    : ((a>=b?'+':'−')+f(Math.abs(a-b)));
  tb+='<tr'+(r===firstVol?' class="sep"':'')+'><td>'+lbl+'</td>'+
   '<td class="r">'+(has?f(b):'—')+'</td><td class="r">'+f(a)+'</td>'+
   '<td class="r '+(has?cls(dp):'')+'">'+dCell+'</td>'+
   '<td class="r '+(has?cls(dp):'')+'">'+(has?sgn(dp)+'%':'—')+'</td></tr>';
 });

 // classify: rate move vs volume move
 const rateD=[A.revPerOpv&&B&&pct(A.revPerOpv,B.revPerOpv),
              A.arpob&&B&&B.arpob&&pct(A.arpob,B.arpob),
              A.revPerDisch&&B&&pct(A.revPerDisch,B.revPerDisch)].filter(v=>typeof v==='number'&&isFinite(v));
 const volD=[B&&pct(A.opvPerDay,B.opvPerDay),
             (A.occ&&B&&B.occ)&&pct(A.occ,B.occ),
             B&&pct(A.ipdPerDay,B.ipdPerDay)].filter(v=>typeof v==='number'&&isFinite(v));
 const avg=a=>a.length? a.reduce((x,y)=>x+y,0)/a.length:0;
 const rAvg=avg(rateD), vAvg=avg(volD);
 let diag;
 if(!B) diag='No prior month with data is available for comparison.';
 else if(Math.abs(rAvg)<3&&Math.abs(vAvg)>=4)
   diag='<b>'+(vAvg<0?'Volume, not price.':'Volume-led, not price-led.')+'</b> Realization is holding — the rate '+
     'measures moved '+sgn(rAvg)+'% on average — while the volume measures moved '+sgn(vAvg)+
     '%. Essentially the entire movement is '+(vAvg<0?'fewer visits, fewer discharges and a lighter census':
     'more visits, more discharges and a fuller census')+'.';
 else if(Math.abs(rAvg)>=3&&Math.abs(vAvg)<3)
   diag='<b>Rate and mix, not volume.</b> Throughput is broadly flat (volume measures '+sgn(vAvg)+
     '% on average) while realization moved '+sgn(rAvg)+'% — this is a pricing, payer-mix or case-mix shift.';
 else if(Math.abs(rAvg)>=3&&Math.abs(vAvg)>=3)
   diag='<b>Rate and volume moved together</b> — realization '+sgn(rAvg)+'%, volume '+sgn(vAvg)+
     '%. Both need reading before attributing the gap to either.';
 else diag='<b>Broadly stable month.</b> Realization moved '+sgn(rAvg)+'% and volume '+sgn(vAvg)+
     '% — neither is a material driver at this scale.';

 const lineTxt=B? ('OP '+sgn(pct(A.opPerDay,B.opPerDay))+'%, IP '+sgn(pct(A.ipPerDay,B.ipPerDay))+
   '%, Pharmacy '+sgn(pct(A.phPerDay,B.phPerDay))+'% per day'):'';
 const allSame=B&&[pct(A.opPerDay,B.opPerDay),pct(A.ipPerDay,B.ipPerDay),pct(A.phPerDay,B.phPerDay)]
   .every(v=>Math.sign(v)===Math.sign(dHead));
 let foot=B? ('Revenue lines: '+lineTxt+
   (allSame? ' — all three moved in the same direction, the signature of a throughput shift rather than a mix or rate shift.'
           : ' — the lines diverge, so this is a mix shift rather than a uniform throughput move.')):'';
 if(A.conv!=null&&B&&B.conv!=null&&Math.abs(A.conv-B.conv)>=3)
   foot+=' <b>Watch collections:</b> cash conversion '+(A.conv<B.conv?'slipped':'improved')+' to '+
     A.conv.toFixed(0)+'% of gross from '+B.conv.toFixed(0)+'% in '+B.label+'.';

 let boxes='<div class="vbox"><h3>1 &middot; '+
   (Math.abs(rAvg)<3&&Math.abs(vAvg)>=4? 'It is volume and census, not price':'Rate versus volume')+'</h3>'+
   '<p>'+diag+'</p>'+
   '<table class="vtab"><thead><tr><th>Measure</th><th class="r">'+(B?B.label:'—')+
   '</th><th class="r">'+A.label+'</th><th class="r">&Delta;</th><th class="r">&Delta;%</th></tr></thead><tbody>'+
   tb+'</tbody></table>'+(foot?'<p class="vfoot">'+foot+'</p>':'')+'</div>';

 // ---- box 2: departments -------------------------------------------------
 function movers(aMap,bMap,nA,nB){
  const names=new Set([...Object.keys(aMap),...Object.keys(bMap||{})]);
  const out=[];
  names.forEach(k=>{
   const a=(aMap[k]||0)/nA/L, b=((bMap||{})[k]||0)/nB/L;
   if(a<0.10&&b<0.10) return;
   out.push({k,a,b,d:a-b,p:b? (a/b-1)*100:null});
  });
  return out.sort((x,y)=>x.d-y.d);
 }
 if(B&&Object.keys(A.depts).length&&Object.keys(B.depts).length){
  const mv=movers(A.depts,B.depts,A.n,B.n);
  const dec=mv.filter(m=>m.d<0), gai=mv.filter(m=>m.d>0).sort((x,y)=>y.d-x.d);
  const decTot=dec.reduce((s,m)=>s+m.d,0), gaiTot=gai.reduce((s,m)=>s+m.d,0);
  const net=decTot+gaiTot;
  const top2=dec.slice(0,2).reduce((s,m)=>s+m.d,0);
  const top2Share=net? Math.abs(top2/net)*100:null;
  const row=m=>'<tr><td>'+nmT(m.k)+'</td><td class="r">'+m.b.toFixed(2)+'</td><td class="r">'+
    m.a.toFixed(2)+'</td><td class="r '+(m.d>=0?'good':'bad')+'">'+(m.d>=0?'+':'−')+
    Math.abs(m.d).toFixed(2)+'</td><td class="r '+(m.d>=0?'good':'bad')+'">'+
    (m.p==null?'—':sgn(m.p)+'%')+'</td></tr>';
  const ratio=gaiTot? Math.abs(decTot/gaiTot):null;
  boxes+='<div class="vbox"><h3>2 &middot; '+
   (top2Share!=null&&top2Share>=50? 'The gap is concentrated':'The gap is broad-based')+'</h3>'+
   '<p>Declines total <b>'+fLd(Math.abs(decTot))+'/day</b> against <b>'+fLd(gaiTot)+'/day</b> of gains'+
   (ratio!=null? ' (a '+ratio.toFixed(0)+':1 ratio)':'')+'. The two largest decliners are '+
   (top2Share!=null? '<b>'+top2Share.toFixed(0)+'% of the net movement</b>':'listed first')+
   '. '+dec.length+' departments fell and '+gai.length+' rose.</p>'+
   '<table class="vtab"><thead><tr><th>Department</th><th class="r">'+B.label.slice(0,3)+
   ' ₹L/d</th><th class="r">'+A.label.slice(0,3)+' ₹L/d</th><th class="r">&Delta; ₹L/d</th>'+
   '<th class="r">&Delta;%</th></tr></thead><tbody>'+
   dec.slice(0,10).map(row).join('')+
   (gai.length? '<tr class="sep">'+row(gai[0]).slice(4):'')+
   gai.slice(1,4).map(row).join('')+
   '</tbody></table>'+
   '<p class="vfoot">Doctor-attributed revenue from the Daily MIS doctor sheets, aggregated to '+
   'department and averaged per calendar day; covers '+(A.docRev/A.rev*100).toFixed(1)+
   '% of gross revenue in '+A.label+'. Departments below ₹0.10 L/day in both months are omitted.</p></div>';

  // ---- box 3: doctors ---------------------------------------------------
  const dm={}; Object.entries(A.docs).forEach(([k,v])=>dm[k]=v.rev);
  const bm={}; Object.entries(B.docs).forEach(([k,v])=>bm[k]=v.rev);
  const dv2=movers(dm,bm,A.n,B.n);
  const ddec=dv2.filter(m=>m.d<0), dgai=dv2.filter(m=>m.d>0).sort((x,y)=>y.d-x.d);
  const drow=m=>{const dp=A.docs[m.k]||B.docs[m.k]||{};
   return '<tr><td class="doc">'+nmT(m.k)+'</td><td class="doc">'+nmT(dp.dept||'—')+
    '</td><td class="r">'+m.b.toFixed(2)+'</td><td class="r">'+m.a.toFixed(2)+
    '</td><td class="r '+(m.d>=0?'good':'bad')+'">'+(m.d>=0?'+':'−')+Math.abs(m.d).toFixed(2)+
    '</td><td class="r '+(m.d>=0?'good':'bad')+'">'+(m.p==null?'—':sgn(m.p)+'%')+'</td></tr>';};
  const top2d=ddec.slice(0,2), gain6=dgai.slice(0,6);
  const t2=Math.abs(top2d.reduce((s,m)=>s+m.d,0)), g6=gain6.reduce((s,m)=>s+m.d,0);
  // volume-vs-rate read on the two biggest decliners
  const vr=top2d.map(m=>{
   const a=A.docs[m.k]||{}, b=B.docs[m.k]||{};
   const oa=(a.opv||0)/A.n, ob=(b.opv||0)/B.n, da=(a.dis||0)/A.n, db=(b.dis||0)/B.n;
   const bits=[];
   if(ob>0.2) bits.push('OP visits '+iN(b.opv)+' → '+iN(a.opv)+' ('+sgn(pct(oa,ob))+'%)');
   if(db>0.2) bits.push('discharges '+iN(b.dis)+' → '+iN(a.dis)+' ('+sgn(pct(da,db))+'%)');
   return bits.length? '<b>'+nmT(m.k)+'</b> '+bits.join(', '):null;
  }).filter(Boolean);
  boxes+='<div class="vbox"><h3>3 &middot; Doctor-level movers</h3>'+
   '<p>The two largest decliners account for <b>'+fLd(t2)+'/day</b> between them, while the six '+
   'largest gainers add back <b>'+fLd(g6)+'/day</b>'+
   (t2>g6*1.5? ' — no offsetting cluster large enough to close the gap.':
    g6>t2? ' — the gainers more than offset the decliners.':' — broadly offsetting.')+'</p>'+
   '<table class="vtab"><thead><tr><th>Consultant</th><th>Department</th><th class="r">'+
   B.label.slice(0,3)+' ₹L/d</th><th class="r">'+A.label.slice(0,3)+' ₹L/d</th>'+
   '<th class="r">&Delta; ₹L/d</th><th class="r">&Delta;%</th></tr></thead><tbody>'+
   ddec.slice(0,9).map(drow).join('')+
   (dgai.length? '<tr class="sep">'+drow(dgai[0]).slice(4):'')+
   dgai.slice(1,6).map(drow).join('')+
   '</tbody></table>'+
   (vr.length? '<p class="vfoot">Volume vs rate on the two largest declines: '+vr.join('; ')+
     '. Books below ₹0.10 L/day in both months are omitted.</p>':'')+'</div>';

  // ---- box 4: generated read --------------------------------------------
  let read='<div class="vbox read"><h3>4 &middot; The read</h3>';
  read+='<p><b>'+A.label+' closed at '+fmtCr(A.rev)+' against '+fmtCr(A.bud)+' of budget, '+
   sgn(dBud)+'%'+(dBud<0?' short':' ahead')+'.</b> Per calendar day the month ran '+sgn(dHead)+
   '% against '+B.label+
   (calShare!=null? ', of which roughly '+calShare.toFixed(0)+'% is calendar — '+nwA+
     ' non-working days against '+nwB+' — leaving '+sgn(dWork)+
     '% as underlying weekday movement.':'; the ex-Sunday, ex-holiday run-rate moved '+sgn(dWork)+'%.')+'</p>';
  read+='<p>'+diag.replace(/<\/?b>/g,'')+' '+
   (allSame? 'All three revenue lines moved together ('+lineTxt+'), which points to throughput rather than mix.'
           : 'The revenue lines diverged ('+lineTxt+'), which points to mix rather than throughput.')+' '+
   (top2Share!=null&&top2Share>=50
     ? 'The movement is concentrated: the top two departments are '+top2Share.toFixed(0)+
       '% of the net change, so this reads as a small number of books rather than a system-wide shift.'
     : 'The movement is broad-based: '+dec.length+' departments fell against '+gai.length+
       ' that rose, and the top two are only '+(top2Share!=null?top2Share.toFixed(0):'—')+
       '% of the net change — that argues against a consultant-availability explanation and towards a demand or funnel issue.')+'</p>';
  // generated action list, keyed off what actually moved
  const acts=[];
  if(B&&pct(A.opvPerDay,B.opvPerDay)<-4&&Math.abs(pct(A.revPerOpv,B.revPerOpv))<3)
   acts.push('the OP funnel — split the '+sgn(A.opvPerDay-B.opvPerDay)+
     ' visits/day into walk-in vs referral vs follow-up, since flat realization points squarely here');
  if(ddec.length) acts.push(nmT(ddec[0].k)+'’s book, which alone is '+fLd(Math.abs(ddec[0].d))+
     '/day of the movement');
  if(A.occ&&B&&B.occ&&pct(A.occ,B.occ)<-4)
   acts.push('bed and OT utilisation — occupancy moved '+sgn(A.occ-B.occ)+
     ' points to '+A.occ.toFixed(1)+'%');
  if(B&&pct(A.ipdPerDay,B.ipdPerDay)<-4)
   acts.push('admission and discharge slots — discharges '+sgn(pct(A.ipdPerDay,B.ipdPerDay))+'% per day');
  if(A.conv!=null&&B&&B.conv!=null&&A.conv<B.conv-3)
   acts.push('the collections slip to '+A.conv.toFixed(0)+
     '% cash conversion, which may be a billing-cycle artefact but is worth confirming before close');
  if(acts.length) read+='<p><b>What to test first:</b> '+
   acts.map((a,i)=>'('+(i+1)+') '+a).join('; ')+'.</p>';
  // live-month caveat
  if(Lv) read+='<p class="vwarn">'+Lv.label+' is '+Lv.n+' day'+(Lv.n===1?'':'s')+
   ' in at '+fLd(Lv.perDay/L)+'/day against '+fLd(Lv.bud/Lv.n/L)+'/day of budget ('+
   sgn(pct(Lv.rev,Lv.bud))+'%)'+
   ((Lv.wkDays===Lv.n)
     ? ' — all working days so far, so against '+A.label+'\u2019s ex-Sunday, ex-holiday '+
       fLd(A.wkPerDay/L)+'/day that is '+sgn(pct(Lv.wkPerDay,A.wkPerDay))+'%'
     : '; on the comparable ex-Sunday, ex-holiday basis '+fLd(Lv.wkPerDay/L)+'/day against '+
       A.label+'\u2019s '+fLd(A.wkPerDay/L)+'/day ('+sgn(pct(Lv.wkPerDay,A.wkPerDay))+'%)')+
   '. At this few days the department and consultant '+
   'splits above are not yet meaningful for '+Lv.label+
   ' — the panel rolls forward to it once the month closes.</p>';
  read+='</div>';
  boxes+=read;
 }
 document.getElementById('vznGrid').innerHTML=boxes;

 // ---- compact live-month strip above the panel --------------------------
 if(Lv){
  const lp=pct(Lv.rev,Lv.bud);
  document.getElementById('mtdStrip').innerHTML=
   '<div class="panel vzn" style="padding-bottom:6px">'+
   '<h2>'+Lv.label+' month-to-date — '+Lv.n+' of '+dim(Lv.mk)+' days</h2>'+
   '<div class="note">Live tracker. The variance panel below analyses the last closed month, '+
   'which is where the department and consultant detail is statistically readable.</div>'+
   '<div class="vstrip">'+
    st('MTD revenue',fmtCr(Lv.rev),'vs '+fmtCr(Lv.bud)+' budget · '+sgn(lp)+'%',cls(lp))+
    st('Working-day run-rate',fLd(Lv.wkPerDay/L)+'/day',
      A? (A.label+' '+fLd(A.wkPerDay/L)+'/day · '+sgn(pct(Lv.wkPerDay,A.wkPerDay))+
         '% · ex-Sunday, ex-holiday'):'',A?cls(pct(Lv.wkPerDay,A.wkPerDay)):'')+
    st('OP visits / working day',iN(Lv.wkOpvPerDay),
      A? (A.label+' '+iN(A.wkOpvPerDay)+' · '+sgn(pct(Lv.wkOpvPerDay,A.wkOpvPerDay))+'%'):'',
      A?cls(pct(Lv.wkOpvPerDay,A.wkOpvPerDay)):'')+
    st('Cash conversion',(Lv.conv||0).toFixed(0)+'%',fmtCr(Lv.coll)+' collected',
      A&&A.conv?cls((Lv.conv||0)-A.conv):'')+
   '</div></div>';
 }
})();

// daily chart
const dailyKeys=Object.keys(D.daily).sort();
let dailyChart=null;
window.drawDaily=function(key){
 const rows=D.daily[key];
 if(dailyChart)dailyChart.destroy();
 dailyChart=new Chart(document.getElementById('dailyChart'),{data:{labels:rows.map(r=>r.date.slice(8)+' '+r.dow.slice(0,2)),datasets:[
  {type:'bar',label:'OP',data:rows.map(r=>r.revOP/L),backgroundColor:LT,stack:'a'},
  {type:'bar',label:'IP',data:rows.map(r=>r.revIP/L),backgroundColor:BLUE,stack:'a'},
  {type:'bar',label:'Pharmacy',data:rows.map(r=>r.revPH/L),backgroundColor:MAROON,stack:'a'},
  {type:'line',label:'Budget',data:rows.map(r=>r.budTot/L),borderColor:GRAY,borderDash:[5,4],pointRadius:0,tension:.2}]},
  options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}}},
  scales:{x:{stacked:true,ticks:{font:{size:10}}},y:{stacked:true,title:{display:true,text:'₹ Lakhs'}}}}});
 document.querySelectorAll('.mbtn').forEach(b=>b.classList.toggle('on',b.dataset.k===key));
}
document.getElementById('mbtns').innerHTML=dailyKeys.map(k=>`<button class="mbtn" data-k="${k}" onclick="drawDaily('${k}')">${k}</button>`).join('');
if(dailyKeys.length)drawDaily(dailyKeys[dailyKeys.length-1]);

// monthly chart
const mlabels=cur.months.map(m=>m.month.slice(0,3));
new Chart(document.getElementById('monthlyChart'),{data:{labels:mlabels,datasets:[
 {type:'bar',label:fyKeys[0]+' Actual',data:cur.months.map(m=>m.revTot/CR),backgroundColor:BLUE},
 {type:'bar',label:'Budget',data:cur.months.map(m=>m.budTot/CR),backgroundColor:'rgba(127,140,155,.35)'},
 {type:'line',label:fyKeys[1]+' Actual',data:prev.months.map(m=>m.revTot/CR),borderColor:MAROON,pointRadius:2,tension:.25}]},
 options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}}},scales:{y:{title:{display:true,text:'₹ Cr'}}}}});

// efficiency chart
new Chart(document.getElementById('effChart'),{data:{labels:eff.map(e=>mName(e.mk)),datasets:[
 {type:'bar',label:'Occupancy %',data:eff.map(e=>e.occ),backgroundColor:LT,yAxisID:'y2'},
 {type:'line',label:'ARPOB ₹’000/day',data:eff.map(e=>e.arpob),borderColor:BLUE,pointRadius:3,tension:.25,yAxisID:'y'},
 {type:'line',label:'ALOS (days)',data:eff.map(e=>e.alos),borderColor:MAROON,borderDash:[5,4],pointRadius:3,yAxisID:'y3'}]},
 options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}}},
 scales:{y:{title:{display:true,text:'ARPOB ₹’000'}},y2:{position:'right',grid:{drawOnChartArea:false},max:100,title:{display:true,text:'Occ %'}},y3:{display:false,min:0,max:8}}}});

// collections
const convLine=cur.months.map(m=>m.revTot? m.collTot/m.revTot*100:null);
new Chart(document.getElementById('collChart'),{data:{labels:mlabels,datasets:[
 {type:'bar',label:'Cash',data:cur.months.map(m=>m.collCash/CR),backgroundColor:BLUE,stack:'c'},
 {type:'bar',label:'Credit',data:cur.months.map(m=>m.collCredit/CR),backgroundColor:MAROON,stack:'c'},
 {type:'line',label:'Budgeted collections',data:cur.months.map(m=>m.budCollTot/CR),borderColor:GRAY,borderDash:[5,4],pointRadius:0},
 {type:'line',label:'Coll ÷ Revenue %',data:convLine,borderColor:GOLD,pointRadius:3,yAxisID:'y2'}]},
 options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}}},
 scales:{x:{stacked:true},y:{stacked:true,title:{display:true,text:'₹ Cr'}},y2:{position:'right',grid:{drawOnChartArea:false},min:0,max:120,title:{display:true,text:'%'}}}}});

// volumes
new Chart(document.getElementById('volChart'),{data:{labels:mlabels,datasets:[
 {type:'bar',label:'OP+ER visits',data:cur.months.map(m=>m.opVisits||null),backgroundColor:LT,yAxisID:'y'},
 {type:'line',label:'OP budget',data:cur.months.map(m=>m.budOPv||null),borderColor:GRAY,borderDash:[4,4],pointRadius:0,yAxisID:'y'},
 {type:'line',label:'IP discharges',data:cur.months.map(m=>m.ipDisch||null),borderColor:MAROON,pointRadius:3,yAxisID:'y2'},
 {type:'line',label:'IP budget',data:cur.months.map(m=>m.budIPd||null),borderColor:MAROON,borderDash:[4,4],pointRadius:0,yAxisID:'y2'}]},
 options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}}},
 scales:{y:{title:{display:true,text:'OP visits'}},y2:{position:'right',grid:{drawOnChartArea:false},title:{display:true,text:'IP discharges'}}}}});

// ---------------- doctor league table ----------------
let rowsL=[];
if(hCur){
 const docsC=hCur.doctors, docsP=hPrev? hPrev.doctors:{};
 const totC=Object.values(docsC).reduce((a,v)=>a+(v.rev||0),0);
 const rankOf=docs=>{const o={};Object.entries(docs).sort((a,b)=>(b[1].rev||0)-(a[1].rev||0))
  .forEach(([n],i)=>o[n]=i+1);return o};
 const rkC=rankOf(docsC), rkP=rankOf(docsP);
 const names=new Set([...Object.keys(docsC),...Object.keys(docsP)]);
 names.forEach(n=>{
  const c=docsC[n]||{}, p=docsP[n]||{};
  const dd=D.disByDoc[n]||null;
  const tm=D.docTypeMix[n]||null, oa=D.opAgg[n]||null;
  const revC=(c.rev||0), revP=(p.rev||0);
  const rrC=revC/dCur, rrP=revP/dPrev;
  const st=dd? dd.status:{};
  const tmTot=tm? (tm.IP+tm.OP+tm.PH):0;
  rowsL.push({doc:tc(n),dept:tc(c.dept||p.dept||''),
   revC:revC/L, rrC:rrC/L, revP:revP/L, rrP:rrP/L,
   mom:rrP? (rrC/rrP-1)*100:null,
   share:totC? revC/totC*100:null,
   rkD:(rkC[n]&&rkP[n])? rkP[n]-rkC[n]:null,
   opC:c.opv||0, opP:p.opv||0, admC:c.adm||0, disC:c.dis||0,
   conv:c.opv? (c.adm||0)/c.opv*100:null,
   ipmix:tmTot? tm.IP/tmTot*100:null,
   newp:oa&&oa.tot? oa.new/oa.tot*100:null,
   freep:oa&&oa.tot? oa.free/oa.tot*100:null,
   rpd:c.dis? revC/c.dis/L:null,
   alos:dd&&dd.losN? dd.losSum/dd.losN:null,
   recov:dd&&dd.n? (st['Recovered']||0)/dd.n*100:null,
   dama:dd? (st['DAMA/LAMA']||0):null, exp:dd? (st['Expired']||0):null,
   cash:dd&&dd.n? dd.cash/dd.n*100:null});
 });
 rowsL=rowsL.filter(r=>r.revC>0.01||r.revP>0.01);
 document.getElementById('dlnote').innerHTML=
  `Current = <b>${mName(curM)}</b> (${dCur} days elapsed) vs previous = <b>${prevM?mName(prevM):'—'}</b> (${dPrev} days). `+
  `Revenue/visits/admissions/discharges from the Daily MIS doctor sheets. Conv % = admissions ÷ OP visits (current month). Rank Δ = movement in revenue rank vs previous month. Columns marked * come from the flash detail sheets over ${D.disDates.length} captured days: IP mix (share of doctor revenue billed as IP), New/Free OP visit shares, and from ${D.nDischarges} discharges — status, ALOS and cash mix. Click a header to sort.`;

 // ---------------- auto-detected doctor findings ----------------
 const ddC=hCur.docDaily||{}, ddP=hPrev? (hPrev.docDaily||{}):{};
 const [fy,fm]=curM.split('-').map(Number), fmDays=new Date(fy,fm,0).getDate();
 const dstats=(o,nd)=>{const v=[];for(let i=1;i<=nd;i++)v.push(((o&&o[String(i)])||0)/L);
  const n=v.length,mean=v.reduce((a,b)=>a+b,0)/n;
  const sd=Math.sqrt(v.reduce((a,b)=>a+(b-mean)*(b-mean),0)/Math.max(n-1,1));
  return{n,mean,sd}};
 const F=[]; let screened=0;
 const fmtL=x=>'₹'+Math.abs(x).toFixed(1)+' L';
 const totP=Object.values(docsP).reduce((a,v)=>a+(v.rev||0),0);
 names.forEach(n=>{
  const c=docsC[n]||{}, p=docsP[n]||{};
  const rrCd=(c.rev||0)/L/dCur, rrPd=(p.rev||0)/L/dPrev;
  if(Math.max(rrCd,rrPd)<0.3) return;
  screened++;
  const disp=tc(n), dept=tc(c.dept||p.dept||'');
  // 1. appeared / gone quiet
  if(!(p.rev)&&c.rev&&rrCd>=0.5){
   F.push({cls:'g',doc:disp,dept,imp:rrCd*fmDays,
    claim:`No revenue in ${mName(prevM)}, now running ₹${rrCd.toFixed(2)} L/day — new, returned from leave, or newly attributed.`,
    evid:`₹${((c.rev||0)/L).toFixed(1)} L in ${dCur} days · ${c.opv||0} OP visits · ${c.dis||0} discharges`});
   return;}
  if(!(c.rev)&&p.rev&&rrPd>=0.5){
   F.push({cls:'r',doc:disp,dept,imp:-rrPd*fmDays,
    claim:`Zero revenue so far in ${mName(curM)} after ₹${rrPd.toFixed(2)} L/day in ${mName(prevM)} — leave, exit or attribution gap. Verify.`,
    evid:`prev month: ₹${((p.rev||0)/L).toFixed(1)} L · ${p.opv||0} OP visits`});
   return;}
  // 2. significant run-rate shift, tested vs the doctor's own daily variance
  if(c.rev&&p.rev&&hPrev){
   const sC=dstats(ddC[n],dCur), sP=dstats(ddP[n],dPrev);
   const se=Math.sqrt(sC.sd*sC.sd/sC.n+sP.sd*sP.sd/sP.n);
   const dlt=sC.mean-sP.mean, z=se>0? dlt/se:0;
   if(Math.abs(z)>=2&&Math.abs(dlt)>=0.25&&Math.abs(dlt)/Math.max(sP.mean,.01)>=0.15){
    const opD=pct((c.opv||0)/dCur,(p.opv||1)/dPrev), dsD=pct((c.dis||0)/dCur,(p.dis||1)/dPrev);
    const rpC=c.dis? (c.rev||0)/c.dis/L:null, rpP=p.dis? (p.rev||0)/p.dis/L:null;
    const rpD=(rpC&&rpP)? pct(rpC,rpP):null;
    const parts=[];
    if(p.dis||c.dis)parts.push(['IP volume',dsD]);
    if(rpD!=null)parts.push(['₹/discharge',rpD]);
    if(p.opv||c.opv)parts.push(['OP volume',opD]);
    parts.sort((a,b)=>Math.abs(b[1])-Math.abs(a[1]));
    const drvTxt=parts.length? `mainly ${parts[0][0]} (${parts[0][1]>=0?'+':''}${parts[0][1].toFixed(0)}%)`+
      (parts[1]? `; ${parts[1][0]} ${parts[1][1]>=0?'+':''}${parts[1][1].toFixed(0)}%`:''):'driver unclear';
    F.push({cls:dlt>0?'g':'r',doc:disp,dept,imp:dlt*fmDays,
     claim:`Run-rate ${dlt>0?'up':'down'}: ₹${sP.mean.toFixed(2)} → ₹${sC.mean.toFixed(2)} L/day vs ${mName(prevM)} — ${drvTxt}.`,
     evid:`z = ${z.toFixed(1)} vs own daily variance · ${dCur} vs ${dPrev} days observed`});
    return;}
  }
  // 3. conversion shift on steady OP base
  if((c.opv||0)>=60&&(p.opv||0)>=60){
   const cvC=(c.adm||0)/c.opv*100, cvP=(p.adm||0)/p.opv*100;
   const rp=((c.dis? (c.rev||0)/c.dis: p.dis? (p.rev||0)/p.dis:0))/L;
   if(Math.abs(cvC-cvP)>=3&&Math.abs(cvC-cvP)/Math.max(cvP,1)>=0.25&&rp>0.3){
    const impM=(cvC-cvP)/100*(c.opv/dCur)*fmDays*rp;
    F.push({cls:cvC>cvP?'g':'r',doc:disp,dept,imp:impM,
     claim:`OP→IP conversion ${cvC>cvP?'rose':'fell'} ${cvP.toFixed(1)}% → ${cvC.toFixed(1)}% on a steady OP base.`,
     evid:`OP ${c.opv} (prev ${p.opv}) · adm ${c.adm||0} (prev ${p.adm||0}) · @ ₹${rp.toFixed(2)} L/disch`});
    return;}
  }
 });
 // 4. portfolio concentration
 if(totC&&totP){
  const top5=docs=>Object.values(docs).map(v=>v.rev||0).sort((a,b)=>b-a).slice(0,5).reduce((a,b)=>a+b,0);
  const shC=top5(docsC)/totC*100, shP=top5(docsP)/totP*100;
  if(Math.abs(shC-shP)>=2)
   F.push({cls:'y',doc:'Portfolio',dept:'concentration',imp:0,
    claim:`Top-5 doctors now ${shC.toFixed(0)}% of doctor-attributed revenue (${shP.toFixed(0)}% in ${mName(prevM)}) — key-person risk ${shC>shP?'rising':'easing'}.`,
    evid:`${Object.keys(docsC).length} active doctors this month`});
 }
 F.sort((a,b)=>Math.abs(b.imp)-Math.abs(a.imp));
 const FK=F.slice(0,9);
 document.getElementById('findings').innerHTML=FK.length? FK.map(c=>
  `<div class="fcard ${c.cls}"><div class="who">${c.doc}<span class="dp">${c.dept}</span></div>`+
  `<div class="claim">${c.claim}</div>`+
  (c.imp?`<div class="imp">${c.imp>0?'+':'−'}${fmtL(c.imp)} / month</div>`:'')+
  `<div class="evid">${c.evid}</div></div>`).join(''):
  '<div class="note">No statistically significant doctor-level shifts detected vs the previous month.</div>';
 document.getElementById('findnote').innerHTML=
  `Hypothesis screen: each doctor's daily revenue in <b>${mName(curM)}</b> (${dCur} days) is tested against <b>${prevM?mName(prevM):'—'}</b> using their own day-to-day variance (Welch z ≥ 2, shift ≥ ₹0.25 L/day and ≥ 15%). ${screened} doctors screened → ${F.length} findings; top ${FK.length} shown, ranked by monthly ₹ impact. Early-month results (&lt; 10 days) carry wider uncertainty.`;
}
let sortK='revC',sortDir=-1;
function renderLeague(){
 const q=(document.getElementById('docFilter').value||'').toUpperCase();
 let rows=rowsL.filter(r=>!q||r.doc.toUpperCase().includes(q)||r.dept.toUpperCase().includes(q));
 rows.sort((a,b)=>{const x=a[sortK],y=b[sortK];
  if(x==null&&y==null)return 0; if(x==null)return 1; if(y==null)return -1;
  return (x<y?-1:x>y?1:0)*(typeof x==='string'?-sortDir:sortDir);});
 const f=(v,d=1)=>v==null?'—':v.toFixed(d);
 const momCell=v=>v==null?'—':`<span class="tag ${v>=10?'g':v<=-10?'r':'y'}">${v>=0?'+':''}${v.toFixed(0)}%</span>`;
 document.querySelector('#league tbody').innerHTML=rows.slice(0,60).map(r=>
  `<tr><td class="doc" title="${r.doc}">${r.doc}</td><td class="doc">${r.dept}</td>`+
  `<td class="r"><b>${f(r.revC)}</b></td><td class="r">${f(r.rrC,2)}</td>`+
  `<td class="r">${f(r.revP)}</td><td class="r">${f(r.rrP,2)}</td>`+
  `<td class="r">${momCell(r.mom)}</td>`+
  `<td class="r">${r.share==null?'—':r.share.toFixed(1)+'%'}</td>`+
  `<td class="r">${r.rkD==null?'—':(r.rkD>0?'▲'+r.rkD:r.rkD<0?'▼'+(-r.rkD):'=')}</td>`+
  `<td class="r">${r.opC||'—'}</td><td class="r">${r.opP||'—'}</td>`+
  `<td class="r">${r.admC||'—'}</td><td class="r">${r.disC||'—'}</td>`+
  `<td class="r">${r.conv==null?'—':r.conv.toFixed(1)+'%'}</td>`+
  `<td class="r">${r.ipmix==null?'—':r.ipmix.toFixed(0)+'%'}</td>`+
  `<td class="r">${r.newp==null?'—':r.newp.toFixed(0)+'%'}</td>`+
  `<td class="r">${r.freep==null?'—':r.freep.toFixed(0)+'%'}</td>`+
  `<td class="r">${f(r.rpd,2)}</td><td class="r">${f(r.alos)}</td>`+
  `<td class="r">${r.recov==null?'—':r.recov.toFixed(0)+'%'}</td>`+
  `<td class="r">${r.dama==null?'—':r.dama}</td><td class="r">${r.exp==null?'—':r.exp}</td>`+
  `<td class="r">${r.cash==null?'—':r.cash.toFixed(0)+'%'}</td></tr>`).join('');
}
document.querySelectorAll('#league th').forEach(th=>th.onclick=()=>{
 const k=th.dataset.k; if(sortK===k)sortDir*=-1; else {sortK=k;sortDir=-1;} renderLeague();});
document.getElementById('docFilter').oninput=renderLeague;
renderLeague();

// ---------------- doctor x day revenue matrix ----------------
const mxMonths=hMonths.filter(mk=>D.history[mk].docDaily&&Object.keys(D.history[mk].docDaily).length);
let mxMonth=mxMonths[mxMonths.length-1]||null;
window.drawMatrix=function(mk){
 mxMonth=mk;
 const h=D.history[mk], dd=h.docDaily||{};
 const nDays=daysIn(mk,h);
 const q=(document.getElementById('mxFilter').value||'').toUpperCase();
 // rows sorted by month total desc
 let docs=Object.keys(dd).map(n=>{
  const tot=Object.values(dd[n]).reduce((a,v)=>a+v,0);
  return {n, tot};}).filter(x=>x.tot>1000)
  .filter(x=>!q||x.n.includes(q))
  .sort((a,b)=>b.tot-a.tot).slice(0,50);
 const days=[...Array(nDays).keys()].map(i=>i+1);
 const [y,m]=mk.split('-').map(Number);
 const dows=days.map(d=>'SMTWTFS'[new Date(y,m-1,d).getDay()]);
 document.querySelector('#matrix thead').innerHTML=
  '<tr><th style="position:sticky;left:0;background:#fff;z-index:2">Doctor</th>'+
  days.map((d,i)=>`<th class="r" style="min-width:44px${dows[i]==='S'?';color:#c0392b':''}">${d}<br><span style="font-weight:400">${dows[i]}</span></th>`).join('')+
  '<th class="r" style="position:sticky;right:0;background:#fff;z-index:2;border-left:2px solid #e3e8ee">Total ₹L</th></tr>';
 const maxAll=Math.max(...docs.slice(0,15).map(x=>Math.max(...Object.values(dd[x.n]))));
 document.querySelector('#matrix tbody').innerHTML=docs.map(x=>{
  const row=dd[x.n];
  return '<tr><td class="doc" style="position:sticky;left:0;background:#fff;z-index:1" title="'+tc(x.n)+'">'+tc(x.n)+'</td>'+
   days.map(d=>{
    const v=row[String(d)]||0;
    if(!v) return '<td class="r" style="color:#c8d0d9">·</td>';
    const a=Math.min(v/maxAll,1);
    const bg=`rgba(43,124,190,${(0.06+a*0.5).toFixed(2)})`;
    const txt=v>=1e5? (v/L).toFixed(1) : (v/1000).toFixed(0)+'k';
    return `<td class="r" style="background:${bg}${a>0.65?';color:#fff':''}" title="₹${Math.round(v).toLocaleString('en-IN')}">${txt}</td>`;
   }).join('')+
   `<td class="r" style="position:sticky;right:0;background:#fff;z-index:1;border-left:2px solid #e3e8ee"><b>${(x.tot/L).toFixed(1)}</b></td></tr>`;
 }).join('');
 document.getElementById('mxnote').innerHTML=
  `Billed gross revenue attributed to each doctor, by calendar day of <b>${mName(mk)}</b> (₹ Lakhs; values under ₹1 L shown as ₹’000 with “k”). Sundays in red. Top 50 doctors — use the filter for others. Note: this is billed revenue, not cash collected — the MIS does not attribute collections to doctors.`;
 document.querySelectorAll('#mxbtns .mbtn').forEach(b=>b.classList.toggle('on',b.dataset.k===mk));
}
document.getElementById('mxbtns').innerHTML=mxMonths.map(k=>`<button class="mbtn" data-k="${k}" onclick="drawMatrix('${k}')">${k}</button>`).join('');
document.getElementById('mxFilter').oninput=()=>drawMatrix(mxMonth);
if(mxMonth)drawMatrix(mxMonth);

// ---------------- department league table ----------------
if(hCur){
 const agg=(h,days)=>{const m={};Object.values(h.doctors).forEach(v=>{
  const k=v.dept||'UNALLOCATED';
  const r=m[k]=m[k]||{rev:0,opv:0,adm:0,dis:0};
  r.rev+=(v.rev||0);r.opv+=(v.opv||0);r.adm+=(v.adm||0);r.dis+=(v.dis||0);});return m};
 const aC=agg(hCur,dCur), aP=hPrev? agg(hPrev,dPrev):{};
 // last-year cuts from persisted history (drop last year's end-of-month Daily MIS files to populate)
 const [cy,cmm]=curM.split('-').map(Number);
 const lyKey=(cy-1)+'-'+String(cmm).padStart(2,'0');
 const aLY=D.history[lyKey]? agg(D.history[lyKey],1):{};
 const fyStart=cmm>=4? cy: cy-1;
 const lyFYmonths=hMonths.filter(k=>{const [y,m]=k.split('-').map(Number);return (m>=4? y: y-1)===fyStart-1;});
 const aFY={};
 lyFYmonths.forEach(k=>{const a=agg(D.history[k],1);
  Object.entries(a).forEach(([d,v])=>{aFY[d]=(aFY[d]||0)+v.rev;});});
 // dept-level discharge stats via doctor->dept map
 const d2d={};Object.entries(hCur.doctors).forEach(([n,v])=>{if(v.dept)d2d[n]=v.dept});
 const dStats={};
 Object.entries(D.disByDoc).forEach(([n,v])=>{
  const k=d2d[n]; if(!k)return;
  const s=dStats[k]=dStats[k]||{n:0,losSum:0,losN:0,cash:0};
  s.n+=v.n;s.losSum+=v.losSum;s.losN+=v.losN;s.cash+=v.cash;});
 const totC=Object.values(aC).reduce((a,v)=>a+v.rev,0);
 let rowsD=Object.keys(aC).map(k=>{
  const c=aC[k],p=aP[k]||{rev:0};
  const rrC=c.rev/dCur, rrP=p.rev? p.rev/dPrev:0;
  const s=dStats[k];
  const alos=s&&s.losN? s.losSum/s.losN:null;
  return {dept:tc(k),revC:c.rev/L,rrC:rrC/L,revP:(p.rev||0)/L,rrP:rrP/L,
   revLY:aLY[k]? aLY[k].rev/L:null, revLYfy:aFY[k]!=null? aFY[k]/L:null,
   mom:rrP? (rrC/rrP-1)*100:null, share:totC? c.rev/totC*100:null,
   opC:c.opv,admC:c.adm,disC:c.dis,
   conv:c.opv? c.adm/c.opv*100:null,
   rpd:c.dis? c.rev/c.dis/L:null,
   alos:alos,
   arpob:(alos&&c.dis)? c.rev/(c.dis*alos)/1000:null,
   cash:s&&s.n? s.cash/s.n*100:null};
 }).filter(r=>r.revC>0.5);
 document.getElementById('dtnote').innerHTML=
  `Current = <b>${mName(curM)}</b> (${dCur} days) vs <b>${prevM?mName(prevM):'—'}</b> (${dPrev} days), doctor-attributed revenue rolled up to department. `+
  `* from flash discharge lists (captured days only). † ARPOB proxy = revenue ÷ (discharges × ALOS*) — departmental bed-days are not reported, so treat as directional. `+
  `LY columns need last year's end-of-month Daily MIS files in the folder (parsed once, then remembered) — “—” means not on hand yet. Click headers to sort.`;
 let sK='revC',sD=-1;
 const fmt=(v,d=1)=>v==null?'—':v.toFixed(d);
 function renderDept(){
  rowsD.sort((a,b)=>{const x=a[sK],y=b[sK];
   if(x==null&&y==null)return 0; if(x==null)return 1; if(y==null)return -1;
   return (x<y?-1:x>y?1:0)*(typeof x==='string'?-sD:sD);});
  const momCell=v=>v==null?'—':`<span class="tag ${v>=10?'g':v<=-10?'r':'y'}">${v>=0?'+':''}${v.toFixed(0)}%</span>`;
  document.querySelector('#deptTable tbody').innerHTML=rowsD.map(r=>
   `<tr><td class="doc">${r.dept}</td><td class="r"><b>${fmt(r.revC)}</b></td><td class="r">${fmt(r.rrC,2)}</td>`+
   `<td class="r">${fmt(r.revP)}</td><td class="r">${fmt(r.rrP,2)}</td>`+
   `<td class="r">${fmt(r.revLY)}</td><td class="r">${fmt(r.revLYfy)}</td>`+
   `<td class="r">${momCell(r.mom)}</td><td class="r">${r.share==null?'—':r.share.toFixed(1)+'%'}</td>`+
   `<td class="r">${r.opC||'—'}</td><td class="r">${r.admC||'—'}</td><td class="r">${r.disC||'—'}</td>`+
   `<td class="r">${r.conv==null?'—':r.conv.toFixed(1)+'%'}</td>`+
   `<td class="r">${fmt(r.rpd,2)}</td><td class="r">${fmt(r.alos)}</td>`+
   `<td class="r">${fmt(r.arpob)}</td><td class="r">${r.cash==null?'—':r.cash.toFixed(0)+'%'}</td></tr>`).join('');
 }
 document.querySelectorAll('#deptTable th').forEach(th=>th.onclick=()=>{
  const k=th.dataset.k; if(sK===k)sD*=-1; else {sK=k;sD=-1;} renderDept();});
 renderDept();
}

// doctor concentration (current month)
if(hCur){
 const arr=Object.entries(hCur.doctors).map(([n,v])=>({n:tc(n),r:v.rev||0}))
  .filter(x=>x.r>0).sort((a,b)=>b.r-a.r);
 const tot=arr.reduce((a,x)=>a+x.r,0);
 let cum=0; const top=arr.slice(0,20).map(x=>{cum+=x.r;return {n:x.n,r:x.r,c:cum/tot*100}});
 const t5=arr.slice(0,5).reduce((a,x)=>a+x.r,0)/tot*100, t10=arr.slice(0,10).reduce((a,x)=>a+x.r,0)/tot*100;
 document.getElementById('concnote').textContent=
  `${mName(curM)}: top-5 doctors = ${t5.toFixed(0)}% and top-10 = ${t10.toFixed(0)}% of doctor-attributed revenue (${arr.length} active doctors). Key-person concentration watch.`;
 new Chart(document.getElementById('concChart'),{data:{labels:top.map(x=>x.n),datasets:[
  {type:'bar',label:'Revenue ₹L',data:top.map(x=>x.r/L),backgroundColor:BLUE,yAxisID:'y'},
  {type:'line',label:'Cumulative %',data:top.map(x=>x.c),borderColor:MAROON,pointRadius:2,yAxisID:'y2'}]},
  options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}}},
  scales:{x:{ticks:{font:{size:9},maxRotation:75,minRotation:60}},y:{title:{display:true,text:'₹ L'}},
  y2:{position:'right',min:0,max:100,grid:{drawOnChartArea:false},title:{display:true,text:'Cum %'}}}}});
}

// dept MoM run-rate
if(hCur){
 const agg=(h,days)=>{const m={};Object.values(h.doctors).forEach(v=>{if(v.dept)m[v.dept]=(m[v.dept]||0)+(v.rev||0)/days});return m};
 const a=agg(hCur,dCur), b=hPrev? agg(hPrev,dPrev):{};
 const keys=Object.keys(a).sort((x,y)=>a[y]-a[x]).slice(0,12);
 new Chart(document.getElementById('deptMoM'),{type:'bar',data:{labels:keys.map(tc),datasets:[
  {label:mName(curM)+' ₹L/day',data:keys.map(k=>(a[k]||0)/L),backgroundColor:BLUE},
  {label:(prevM?mName(prevM):'prev')+' ₹L/day',data:keys.map(k=>(b[k]||0)/L),backgroundColor:'rgba(139,26,74,.55)'}]},
  options:{indexAxis:'y',plugins:{legend:{labels:{boxWidth:12,font:{size:11}}}},
  scales:{x:{title:{display:true,text:'₹ Lakhs / day'}},y:{ticks:{font:{size:10.5}}}}}});
}

// discharge outcomes + payer
document.getElementById('disnote').textContent=
 `${D.nDischarges} discharges across ${D.disDates.length} captured flash days (${D.disDates[0]||''} → ${D.disDates[D.disDates.length-1]||''}). Left: clinical status. Right: payer scheme mix.`;
const sm=D.statusMix, pm=D.payerMix;
new Chart(document.getElementById('statusChart'),{type:'doughnut',
 data:{labels:Object.keys(sm),datasets:[{data:Object.values(sm),backgroundColor:[ '#1a7f4e',MAROON,'#c8952b',BLUE,GRAY]}]},
 options:{plugins:{legend:{position:'bottom',labels:{boxWidth:11,font:{size:10.5}}}}}});
new Chart(document.getElementById('payerChart'),{type:'doughnut',
 data:{labels:Object.keys(pm),datasets:[{data:Object.values(pm),backgroundColor:[BLUE,MAROON,'#c8952b',GRAY]}]},
 options:{plugins:{legend:{position:'bottom',labels:{boxWidth:11,font:{size:10.5}}}}}});

// dept pareto (flash captured days)
document.getElementById('deptnote').textContent='Net service revenue summed over '+D.deptDates.length+' captured day(s).';
new Chart(document.getElementById('deptChart'),{type:'bar',data:{labels:D.deptTop.map(d=>d.name),
 datasets:[{label:'Revenue (₹ L)',data:D.deptTop.map(d=>d.rev/L),backgroundColor:BLUE}]},
 options:{indexAxis:'y',plugins:{legend:{display:false}},scales:{x:{title:{display:true,text:'₹ Lakhs'}},y:{ticks:{font:{size:10.5}}}}}});

document.getElementById('foot').innerHTML='<b>Files parsed:</b> '+D.filesParsed.map(f=>'<span class="pill">'+f+'</span>').join('')+
 '<br>Gross revenue per Daily Revenue Flash. Doctor-month figures from Daily MIS doctor sheets (newest file per month; current month is MTD — compare on ₹/day run-rate). Discharge status / ALOS / payer mix cover only dates with a flash file on hand. ARPOB = gross revenue ÷ occupied bed-days; ALOS from MIS MoM sheet. Operations include VPSLMC (satellite) figures where the source does.'+
 (D.aop? '<br><b>FY 26-27 projection tab:</b> AOP plan from '+D.aop.source+' (\'Monthly P&L\'). Plan line = IP + OP revenue, which is what the flash gross figure covers; the plan\'s F&amp;B (₹'+((D.aop.fyPlanTotal-D.aop.fyPlanRev)/CR).toFixed(1)+' Cr incl. other income) is excluded so the comparison is like-for-like. Projection is a run-rate extrapolation, not a bottom-up build — it holds current realization and payer mix flat.':'');

// ============================ FY 26-27 PROJECTION TAB ============================
const VIEWS={ops:'Ops',proj:'Proj',cmi:'Cmi',fy27:'Fy27'};
function showView(v){
 Object.keys(VIEWS).forEach(k=>{
  const el=document.getElementById('view'+VIEWS[k]), tb=document.getElementById('tab'+VIEWS[k]);
  if(el) el.classList.toggle('on',k===v);
  if(tb) tb.classList.toggle('on',k===v);
 });
 if(v==='proj') setTimeout(drawProj,30);
 if(v==='cmi') setTimeout(window.drawCmi,30);
 if(v==='fy27') setTimeout(window.drawFy27,30);
}

const MO=['April','May','June','July','August','September','October','November','December','January','February','March'];
const MOS=['Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar'];
const fyYm=(i,base)=>[base+(i<9?0:1), i<9? i+4 : i-8];
const dim=(y,m)=>new Date(y,m,0).getDate();
// FY start years, e.g. 'FY 2026-27' -> 2026
const fyBase=k=>{const m=(k||'').match(/(\d{4})/);return m? +m[1]:2026;};
const Y27=fyBase(fyKeys[0]), Y26=fyBase(fyKeys[1]);
const d27=MO.map((_,i)=>dim(...fyYm(i,Y27))), d26=MO.map((_,i)=>dim(...fyYm(i,Y26)));

// current month index + banked daily series
const pjKeys=Object.keys(D.daily).sort();
const pjKey=pjKeys[pjKeys.length-1];
const curIdx=(()=>{const m=+pjKey.split('-')[1];return m>=4? m-4 : m+8;})();
const dayRows=(D.daily[pjKey]||[]).filter(r=>r.revTot>0);
const banked=dayRows.reduce((a,r)=>a+r.revTot,0);
const bankedDays=dayRows.length;

// seasonal index: FY25-26 per-day, normalised to that FY's same-month-as-current
const pd26=MO.map((_,i)=>(prev.months[i]? prev.months[i].revTot:0)/d26[i]);
const seasIdx=pd26.map(v=>pd26[curIdx]? v/pd26[curIdx]:1);

// run-rate windows
const RRW=[{k:'7',lbl:'Last 7 days'},{k:'14',lbl:'Last 14 days'},
           {k:'mtd',lbl:'Full month-to-date'},{k:'q',lbl:'Last 3 months'}];
let rrKey='mtd';
function runRate(){
 if(rrKey==='q'){
  let rev=0,dys=0;
  for(let i=Math.max(0,curIdx-2);i<curIdx;i++){if(cur.months[i]){rev+=cur.months[i].revTot;dys+=d27[i];}}
  rev+=banked;dys+=bankedDays;
  return dys? rev/dys : 0;
 }
 if(rrKey==='mtd') return bankedDays? banked/bankedDays : 0;
 const n=+rrKey, sl=dayRows.slice(-n);
 return sl.length? sl.reduce((a,r)=>a+r.revTot,0)/sl.length : 0;
}

const aop=D.aop, aopRev=aop? aop.months.map(m=>m.rev):null;

function build(){
 const rr=runRate(), ramp=+document.getElementById('rampS').value/100,
       seas=document.getElementById('seasS').checked;
 const rows=MO.map((mn,i)=>{
  let rev,basis;
  if(i<curIdx){ rev=cur.months[i]? cur.months[i].revTot:0; basis='Actual'; }
  else if(i===curIdx){ rev=banked+(d27[i]-bankedDays)*rr;
    basis=bankedDays+' d actual + '+(d27[i]-bankedDays)+' d at run-rate'; }
  else { const st=i-curIdx;
    rev=rr*(seas? seasIdx[i]:1)*d27[i]*Math.pow(1+ramp,st);
    basis='Run-rate'+(seas?' × seasonality':'')+(ramp?' × ramp^'+st:''); }
  return {i:i,mn:mn,ms:MOS[i],days:d27[i],idx:seasIdx[i],rev:rev,basis:basis,
          a26:prev.months[i]? prev.months[i].revTot:0,
          plan:aopRev? aopRev[i]:null, closed:i<curIdx, isCur:i===curIdx};
 });
 return {rows:rows,rr:rr,ramp:ramp,seas:seas,
   tot:rows.reduce((a,r)=>a+r.rev,0),
   t26:rows.reduce((a,r)=>a+r.a26,0),
   plan:aopRev? aopRev.reduce((a,b)=>a+b,0):null};
}
function fyTotalFor(rampPct){
 const rr=runRate(), seas=document.getElementById('seasS').checked, ramp=rampPct/100;
 let t=0;
 for(let i=0;i<12;i++){
  if(i<curIdx) t+=cur.months[i]? cur.months[i].revTot:0;
  else if(i===curIdx) t+=banked+(d27[i]-bankedDays)*rr;
  else t+=rr*(seas? seasIdx[i]:1)*d27[i]*Math.pow(1+ramp,i-curIdx);
 }
 return t;
}

let pCharts={};
function drawProj(){
 const M=build(), rows=M.rows;
 document.getElementById('rrVal').textContent='₹'+(M.rr/CR).toFixed(3)+' Cr/day';
 document.getElementById('rampVal').textContent=(M.ramp>=0?'+':'')+(M.ramp*100).toFixed(2)+'% / month';
 const dv=(v)=>(v>=0?'+₹':'−₹')+Math.abs(v/CR).toFixed(1);
 const cardsHtml=[
  ['FY 26-27 projection',fmtCr(M.tot),'vs FY 25-26 ₹'+(M.t26/CR).toFixed(1)+' Cr',M.tot>=M.t26],
  ['YoY growth',((M.tot/M.t26-1)*100).toFixed(1)+'%','on '+fyKeys[1],M.tot>=M.t26],
  M.plan? ['vs AOP plan',dv(M.tot-M.plan)+' Cr','plan ₹'+(M.plan/CR).toFixed(1)+' Cr (IP+OP)',M.tot>=M.plan]
        : ['AOP plan','n/a','plan file not found',true],
  ['Banked so far',fmtCr(rows.filter(r=>r.closed).reduce((a,r)=>a+r.rev,0)+banked),
    rows.filter(r=>r.closed).length+' closed months + '+bankedDays+' d of '+MOS[curIdx],true],
  ['H2 implied (Oct–Mar)',fmtCr(rows.slice(6).reduce((a,r)=>a+r.rev,0)),
    'vs ₹'+(rows.slice(6).reduce((a,r)=>a+r.a26,0)/CR).toFixed(1)+' Cr LY',true]
 ];
 document.getElementById('pcards').innerHTML=cardsHtml.map((c,n)=>
  '<div class="card'+(n===0?'':' m')+'"><div class="lbl">'+c[0]+'</div><div class="val">'+c[1]+
  '</div><div class="delta '+(c[3]?'up':'dn')+'">'+c[2]+'</div></div>').join('');

 document.getElementById('pchartnote').innerHTML='₹ Cr per month. Solid blue = closed actuals, hatched = projected. '+
  (aop? 'Maroon line = AOP plan (IP + OP). ':'')+'Gray = FY 25-26 actual.';

 const proj=rows.map(r=>r.closed? null : r.rev/CR);
 const act=rows.map(r=>r.closed? r.rev/CR : null);
 if(pCharts.p) pCharts.p.destroy();
 pCharts.p=new Chart(document.getElementById('projChart'),{data:{labels:MOS,datasets:[
  {type:'bar',label:'FY 25-26 actual',data:rows.map(r=>r.a26/CR),backgroundColor:'rgba(127,140,155,.45)',borderRadius:3,order:4},
  {type:'bar',label:'FY 26-27 actual',data:act,backgroundColor:BLUE,borderRadius:3,order:3},
  {type:'bar',label:'FY 26-27 projected',data:proj,backgroundColor:LT,borderColor:BLUE,borderWidth:1.4,borderRadius:3,order:3}
 ].concat(aop? [{type:'line',label:'AOP plan',data:aopRev.map(v=>v/CR),borderColor:MAROON,borderWidth:2,
   pointRadius:3,pointBackgroundColor:MAROON,tension:.25,order:1}]:[])},
  options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}},
   tooltip:{callbacks:{label:c=>c.dataset.label+': ₹'+(+c.raw).toFixed(2)+' Cr'}}},
   scales:{y:{title:{display:true,text:'₹ Cr'},beginAtZero:true},x:{stacked:false}}}});

 // cumulative
 let ca=0,cp=0; const cumP=[],cumPlan=[];
 rows.forEach(r=>{ca+=r.rev;cumP.push(ca/CR);if(aop){cp+=r.plan;cumPlan.push(cp/CR);}});
 if(pCharts.c) pCharts.c.destroy();
 pCharts.c=new Chart(document.getElementById('cumChart'),{type:'line',data:{labels:MOS,datasets:[
  {label:'Projection cumulative',data:cumP,borderColor:BLUE,backgroundColor:'rgba(43,124,190,.14)',fill:aop?'+1':'origin',borderWidth:2,pointRadius:2.5,tension:.2}
 ].concat(aop? [{label:'AOP plan cumulative',data:cumPlan,borderColor:MAROON,borderWidth:2,borderDash:[6,4],pointRadius:2.5,tension:.2,fill:false}]:[])},
  options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}},
   tooltip:{callbacks:{label:c=>c.dataset.label+': ₹'+(+c.raw).toFixed(1)+' Cr'}}},
   scales:{y:{title:{display:true,text:'₹ Cr cumulative'}}}}});

 // scenario fan
 const fanX=[];for(let p=-2;p<=5;p+=0.5)fanX.push(p);
 const fanY=fanX.map(p=>fyTotalFor(p)/CR);
 const here=fanX.map(p=>Math.abs(p-M.ramp*100)<0.26? fyTotalFor(p)/CR : null);
 if(pCharts.f) pCharts.f.destroy();
 pCharts.f=new Chart(document.getElementById('fanChart'),{data:{labels:fanX.map(p=>(p>=0?'+':'')+p+'%'),datasets:[
  {type:'line',label:'FY total',data:fanY,borderColor:BLUE,borderWidth:2,pointRadius:2,tension:.2,fill:false},
  {type:'line',label:'Current setting',data:here,borderColor:MAROON,pointRadius:6,pointBackgroundColor:MAROON,showLine:false}
 ].concat(aop? [{type:'line',label:'AOP plan',data:fanX.map(()=>M.plan/CR),borderColor:GRAY,borderWidth:1.4,borderDash:[5,4],pointRadius:0}]:[])},
  options:{plugins:{legend:{labels:{boxWidth:12,font:{size:11}}},
   tooltip:{callbacks:{label:c=>c.dataset.label+': ₹'+(+c.raw).toFixed(1)+' Cr'}}},
   scales:{y:{title:{display:true,text:'₹ Cr FY total'}},x:{title:{display:true,text:'monthly ramp'}}}}});

 // table
 document.getElementById('ptabnote').innerHTML='Seasonality index = each FY 25-26 month\'s ₹/day relative to '+MOS[curIdx]+
  ' FY 25-26. Shaded rows are closed actuals.'+(aop? ' AOP plan is the IP + OP line from '+aop.source+'.':'');
 const vtag=(v)=>v==null? '—' : '<span class="tag '+(v>=0?'g':'r')+'">'+(v>=0?'+':'')+(v/CR).toFixed(2)+'</span>';
 document.querySelector('#projTable tbody').innerHTML=rows.map(r=>
  '<tr class="'+(r.closed?'pastm':'')+'"><td>'+r.mn+(r.isCur?' <span class="tag y">current</span>':'')+'</td>'+
  '<td class="r">'+r.days+'</td><td class="r">'+r.idx.toFixed(3)+'</td>'+
  '<td class="r">'+(r.a26/CR).toFixed(2)+'</td>'+
  '<td class="r">'+(r.plan!=null?(r.plan/CR).toFixed(2):'—')+'</td>'+
  '<td class="r"><b>'+(r.rev/CR).toFixed(2)+'</b></td>'+
  '<td class="r">'+(r.rev/r.days/CR).toFixed(3)+'</td>'+
  '<td class="r">'+vtag(r.plan!=null? r.rev-r.plan:null)+'</td>'+
  '<td class="r">'+vtag(r.rev-r.a26)+'</td><td style="font-size:11px;color:#7F8C9B">'+r.basis+'</td></tr>').join('')+
  '<tr class="fytot"><td>FY 26-27 total</td><td class="r">'+d27.reduce((a,b)=>a+b,0)+'</td><td class="r"></td>'+
  '<td class="r">'+(M.t26/CR).toFixed(1)+'</td><td class="r">'+(M.plan!=null?(M.plan/CR).toFixed(1):'—')+'</td>'+
  '<td class="r">'+(M.tot/CR).toFixed(1)+'</td><td class="r"></td>'+
  '<td class="r">'+vtag(M.plan!=null? M.tot-M.plan:null)+'</td><td class="r">'+vtag(M.tot-M.t26)+'</td><td></td></tr>';

 // commentary
 const fut=rows.filter(r=>!r.closed&&!r.isCur);
 const wk=fut.slice().sort((a,b)=>(a.rev-a.plan)-(b.rev-b.plan))[0];
 const bst=fut.slice().sort((a,b)=>(b.rev-b.plan)-(a.rev-a.plan))[0];
 const need=M.plan!=null? (M.plan-rows.filter(r=>r.closed||r.isCur).reduce((a,r)=>a+r.rev,0))/
   fut.reduce((a,r)=>a+r.days,0):null;
 document.getElementById('pcomment').innerHTML=
  'At a <b>₹'+(M.rr/CR).toFixed(3)+' Cr/day</b> run-rate ('+RRW.find(w=>w.k===rrKey).lbl.toLowerCase()+')'+
  (M.seas?', FY 25-26 seasonality applied':', flat across months')+
  ' and a <b>'+(M.ramp>=0?'+':'')+(M.ramp*100).toFixed(2)+'%</b> monthly ramp, FY 26-27 lands at <b>'+fmtCr(M.tot)+
  '</b> — '+(M.tot>=M.t26?'up':'down')+' '+Math.abs((M.tot/M.t26-1)*100).toFixed(1)+'% on FY 25-26'+
  (M.plan!=null? ' and ₹'+Math.abs((M.tot-M.plan)/CR).toFixed(1)+' Cr '+(M.tot>=M.plan?'above':'below')+' the AOP plan':'')+'. '+
  (need!=null? 'To land exactly on plan, the remaining '+fut.reduce((a,r)=>a+r.days,0)+' days after '+MOS[curIdx]+
   ' would need to average <b>₹'+(need/CR).toFixed(3)+' Cr/day</b> ('+
   ((need/M.rr-1)*100>=0?'+':'')+((need/M.rr-1)*100).toFixed(0)+'% vs the current run-rate). ':'')+
  (aop? 'The tightest month against plan is <b>'+wk.mn+'</b> (₹'+((wk.rev-wk.plan)/CR).toFixed(2)+
   ' Cr) and the loosest <b>'+bst.mn+'</b> (₹'+((bst.rev-bst.plan)/CR).toFixed(2)+' Cr). ':'')+
  '<span style="color:#7F8C9B">Caveat: this is a top-down run-rate extrapolation. It holds realization, payer mix and case mix at current levels, and '+
  (M.seas?'borrows its monthly shape from last year — so any tariff revision, bed addition or Onam/festival shift moves it independently of days.':
   'applies no seasonality, so it will overstate the Nov trough and understate the Jan–Feb peak.')+'</span>';
}

(function initProj(){
 document.getElementById('rrBtns').innerHTML=RRW.map(w=>
  '<button class="rbtn'+(w.k===rrKey?' on':'')+'" data-w="'+w.k+'">'+w.lbl+'</button>').join('');
 document.querySelectorAll('#rrBtns .rbtn').forEach(b=>b.onclick=()=>{
  rrKey=b.dataset.w;
  document.querySelectorAll('#rrBtns .rbtn').forEach(x=>x.classList.toggle('on',x===b));
  drawProj();});
 document.getElementById('rampS').oninput=drawProj;
 document.getElementById('seasS').onchange=drawProj;
})();

// ============================ CASE MIX (CMI) TAB ============================
// CMI proxy = specialty IP-attributed revenue per discharge, indexed to the
// hospital average over admitting specialties only. No DRG weights exist in the
// source, so acuity is inferred from billed intensity per case.
(function(){
 const BAD=/^(GRAND TOTAL|TOTAL|CHEMOTHERAPY|EXECUTIVE HEALTH)$/i;
 const MINDIS_DEPT=10;
 let minDoc=5;

 // per-doctor IP-attributed revenue: IP service revenue + pharmacy apportioned
 // by that doctor's IP share of service revenue
 function docIp(dn,rec){
  const t=D.docTypeMix[dn]||{}, ip=t.IP||0, op=t.OP||0, ph=t.PH||0, svc=ip+op;
  const sh=svc>0? ip/svc : (rec.dis>0?1:0);
  return ip+ph*sh;
 }
 function alosOf(dn){const l=D.disByDoc[dn]; return (l&&l.losN)? l.losSum/l.losN : 0;}
 function cashOf(dn){const l=D.disByDoc[dn]; return (l&&l.n)? l.cash/l.n : null;}

 function deptAgg(mk){
  const H=(D.history[mk]||{}).doctors||{}, out={};
  Object.keys(H).forEach(dn=>{
   const r=H[dn]; if(BAD.test((r.dept||'').trim())) return;
   const o=out[r.dept]=out[r.dept]||{dept:r.dept,rev:0,iprev:0,dis:0,bd:0,docs:0};
   o.rev+=r.rev; o.iprev+=docIp(dn,r); o.dis+=r.dis; o.bd+=alosOf(dn)*r.dis; o.docs++;
  });
  return out;
 }

 // ---- fixed relative weights, pooled over every history month (the base period).
 // Weights are frozen, so the hospital CMI moves on case-mix shift, not on price.
 function baseAgg(){
  const out={};
  Object.keys(D.history).forEach(mk=>{
   Object.values(deptAgg(mk)).forEach(d=>{
    const o=out[d.dept]=out[d.dept]||{dept:d.dept,iprev:0,dis:0,bd:0};
    o.iprev+=d.iprev; o.dis+=d.dis; o.bd+=d.bd;});});
  return out;
 }
 function weightSet(noDay){
  const B=baseAgg(), inB=d=>!noDay||(d.dis? d.bd/d.dis:0)>=2;
  const adm=Object.values(B).filter(d=>d.dis>=MINDIS_DEPT);
  const base=adm.filter(inB);
  const TI=base.reduce((a,d)=>a+d.iprev,0), TD=base.reduce((a,d)=>a+d.dis,0);
  const avg=TD? TI/TD:0, w={};
  adm.forEach(d=>{w[d.dept]=avg? (d.iprev/d.dis)/avg : 0;});
  return {w:w,avg:avg,cases:TD,nDept:base.length,months:Object.keys(D.history).sort()};
 }
 // CMI = Σ(case weight) ÷ number of cases
 function cmiOf(agg,W,inBase){
  let sw=0,n=0;
  Object.values(agg).forEach(d=>{
   const wt=W.w[d.dept];
   if(wt==null||d.dis<MINDIS_DEPT||!inBase(d))return;
   sw+=wt*d.dis; n+=d.dis;});
  return {cmi:n? sw/n:0,sw:sw,n:n};
 }

 function model(){
  const A=deptAgg(curM), P=prevM? deptAgg(prevM):{};
  const adm=Object.values(A).filter(d=>d.dis>=MINDIS_DEPT);
  const exc=Object.values(A).filter(d=>d.dis<MINDIS_DEPT);
  const noDay=document.getElementById('cmiNoDay').checked;
  const inBase=d=>!noDay || (d.dis? d.bd/d.dis:0)>=2;
  const W=weightSet(noDay);
  const HC=cmiOf(A,W,inBase), HCp=prevM? cmiOf(P,W,inBase):null;
  const base=adm.filter(inBase);
  const TI=base.reduce((a,d)=>a+d.iprev,0), TD=base.reduce((a,d)=>a+d.dis,0),
        TB=base.reduce((a,d)=>a+d.bd,0);
  const avgC=TD? TI/TD:0, avgB=TB? TI/TB:0;
  const pAdm=Object.values(P).filter(d=>d.dis>=MINDIS_DEPT&&inBase(d));
  const pTI=pAdm.reduce((a,d)=>a+d.iprev,0), pTD=pAdm.reduce((a,d)=>a+d.dis,0);
  const rows=adm.map(d=>{
   const al=d.dis? d.bd/d.dis:0;
   const rpc=d.iprev/d.dis, rpb=(d.bd&&al>=1)? d.iprev/d.bd:0, p=P[d.dept];
   return {dept:d.dept,iprev:d.iprev,dis:d.dis,docs:d.docs,inBase:inBase(d),
     alos:al,rpc:rpc,cmi:(W.w[d.dept]!=null? W.w[d.dept]:0),cmiCur:avgC? rpc/avgC:0,
     rpb:rpb,ii:(avgB&&rpb)? rpb/avgB:0,
     shRev:TI? d.iprev/TI:0,shDis:TD? d.dis/TD:0,
     mom:(p&&p.dis>=MINDIS_DEPT)? rpc-(p.iprev/p.dis) : null};
  }).sort((a,b)=>b.iprev-a.iprev);
  const docs=[];
  const H=(D.history[curM]||{}).doctors||{};
  // doctor weights pooled over the same base period
  const DB={};
  Object.keys(D.history).forEach(mk=>{
   const H2=(D.history[mk]||{}).doctors||{};
   Object.keys(H2).forEach(dn=>{const r=H2[dn];
    if(BAD.test((r.dept||'').trim()))return;
    const o=DB[dn]=DB[dn]||{ip:0,dis:0}; o.ip+=docIp(dn,r); o.dis+=r.dis;});});
  const docW=dn=>{const b=DB[dn]; return (b&&b.dis&&W.avg)? (b.ip/b.dis)/W.avg : 0;};
  Object.keys(H).forEach(dn=>{
   const r=H[dn]; if(BAD.test((r.dept||'').trim())||r.dis<minDoc) return;
   const ipr=docIp(dn,r), al=alosOf(dn), bd=(al>=1)? al*r.dis : 0;
   docs.push({doc:dn,dept:r.dept,iprev:ipr,dis:r.dis,alos:al,
     rpc:ipr/r.dis,cmi:docW(dn),
     rpb:bd? ipr/bd:0,ii:(avgB&&bd)?(ipr/bd)/avgB:0,cash:cashOf(dn)});
  });
  docs.sort((a,b)=>b.cmi-a.cmi);
  return {rows:rows,docs:docs,exc:exc.sort((a,b)=>b.rev-a.rev),
    excRev:exc.reduce((a,d)=>a+d.rev,0),
    TI:TI,TD:TD,TB:TB,avgC:avgC,avgB:avgB,noDay:noDay,
    nBase:base.length,nAdm:adm.length,
    momAvg:(pTD&&pTI)? avgC/(pTI/pTD)-1 : null,
    W:W,cmiHosp:HC.cmi,cmiSw:HC.sw,cmiN:HC.n,cmiPrev:HCp? HCp.cmi:null,
    grossRev:mtd.revTot||0};
 }

 let ch={},sortK='iprev',sortA=false,dSortK='cmi',dSortA=false;
 function render(){
  const M=model(), rows=M.rows;
  document.getElementById('cmiMethod').innerHTML=
   '<b>CMI = Σ(case weight) ÷ number of cases</b>, the standard construction. What differs from a coded CMI is the source of the weights: '+
   'no DRG or relative-weight field exists in the flash or the Daily MIS, so each case carries its <b>specialty\'s relative weight</b> '+
   'in place of a DRG weight. Weights are set once over the base period ('+M.W.months.map(mName).join(' + ')+', '+
   M.W.cases.toLocaleString('en-IN')+' cases, hospital average ₹'+(M.W.avg/L).toFixed(2)+' L/case = weight 1.00) and then held fixed, '+
   'so the monthly CMI moves when the mix of cases shifts, not when prices or billing move. '+
   'For each doctor, IP-attributed revenue = IP service revenue + pharmacy apportioned by that doctor\'s IP share of service revenue; '+
   'rolled to specialty and divided by discharges to give that specialty\'s ₹/case. Only <b>admitting specialties</b> '+
   '(≥'+MINDIS_DEPT+' discharges) are weighted and counted. Month: <b>'+mName(curM)+'</b>, '+D.history[curM].daysElapsed+' days elapsed. '+
   'Because the weights are billed-intensity rather than coded acuity, read this as a <b>resource-intensity index</b>: it answers '+
   '"is this month\'s case mix tilted toward the heavier specialties?", not "how sick were these patients?". '+
   'ALOS is the discharge-level average from the flash <i>Dis</i> sheets, so it covers only dates with a flash file on hand; ₹/bed-day is suppressed where ALOS is under 1 day. '+
   '<b>The MoM column and the MoM card apply this month\'s IP/OP mix ratios and ALOS to last month\'s revenue and discharges</b>, because the flash detail sheets only cover the current month — so treat the MoM figures as directional on volume and case value, not on mix shift.';

  document.getElementById('cmiBaseN').textContent=M.TD+' cases · '+M.nBase+' of '+M.nAdm+' specialties';
  const top=rows.slice().sort((a,b)=>b.cmi-a.cmi)[0], bot=rows.slice().sort((a,b)=>a.cmi-b.cmi)[0];
  const dCmi=(M.cmiPrev!=null&&M.cmiPrev)? M.cmiHosp-M.cmiPrev:null;
  const cards=[
   ['Hospital CMI ('+mName(curM)+')',M.cmiHosp.toFixed(3),
     'Σw '+M.cmiSw.toFixed(0)+' ÷ '+M.cmiN.toLocaleString('en-IN')+' cases'+
     (dCmi!=null? ' · '+(dCmi>=0?'▲ +':'▼ ')+dCmi.toFixed(3)+' vs '+mName(prevM):''),
     dCmi==null||dCmi>=0],
   ['Hospital ₹ / case','₹'+(M.avgC/L).toFixed(2)+' L',M.momAvg!=null?
     ((M.momAvg>=0?'▲ +':'▼ ')+(M.momAvg*100).toFixed(1)+'% vs '+mName(prevM)):'no prior month',M.momAvg==null||M.momAvg>=0],
   ['Blended ALOS',(M.TB/M.TD).toFixed(2)+' d','MoM sheet ALOS '+((D.momFY[curM]&&D.momFY[curM].alos)?D.momFY[curM].alos.toFixed(2):'—')+' d',true],
   ['₹ / bed-day','₹'+(M.avgB/L).toFixed(2)+' L','across '+Math.round(M.TB).toLocaleString('en-IN')+' bed-days',true],
   ['Highest acuity',top? top.dept.slice(0,20):'—',top? 'CMI '+top.cmi.toFixed(2)+' · '+top.dis+' cases':'',true],
   ['Lowest acuity',bot? bot.dept.slice(0,20):'—',bot? 'CMI '+bot.cmi.toFixed(2)+' · '+bot.dis+' cases':'',false],
   ['In-index coverage',(M.TI/(M.TI+M.excRev)*100).toFixed(0)+'%',
     rows.length+' admitting specialties · '+M.TD+' cases',true]
  ];
  document.getElementById('cmiCards').innerHTML=cards.map((c,n)=>
   '<div class="card'+(n<3?'':' m')+'"><div class="lbl">'+c[0]+'</div><div class="val" style="font-size:'+
   (typeof c[1]==='string'&&c[1].length>12?'15px':'21px')+'">'+c[1]+'</div><div class="delta '+(c[3]?'up':'dn')+'">'+c[2]+'</div></div>').join('');

  // CMI bar chart
  const cs=rows.slice().sort((a,b)=>b.cmi-a.cmi);
  if(ch.b)ch.b.destroy();
  ch.b=new Chart(document.getElementById('cmiChart'),{type:'bar',
   data:{labels:cs.map(r=>tc(r.dept)),datasets:[{data:cs.map(r=>r.cmi),
    backgroundColor:cs.map(r=>r.cmi>=1?BLUE:MAROON),borderRadius:3}]},
   options:{indexAxis:'y',plugins:{legend:{display:false},
    tooltip:{callbacks:{label:c=>'CMI '+(+c.raw).toFixed(2)+' · ₹'+(cs[c.dataIndex].rpc/L).toFixed(2)+' L/case · '+cs[c.dataIndex].dis+' cases'}}},
    scales:{x:{title:{display:true,text:'Relative weight (1.00 = base-period hospital average)'},
      grid:{color:ctx=>Math.abs(ctx.tick.value-1)<0.001?'#243342':'rgba(0,0,0,.06)'}},
     y:{ticks:{font:{size:10}}}}}});

  // scatter
  const mx=Math.max.apply(null,rows.map(r=>r.iprev))||1;
  if(ch.s)ch.s.destroy();
  ch.s=new Chart(document.getElementById('cmiScatter'),{type:'bubble',
   data:{datasets:[{data:rows.map(r=>({x:r.cmi,y:r.dis,r:6+22*Math.sqrt(r.iprev/mx),d:r})),
    backgroundColor:rows.map(r=>r.cmi>=1?'rgba(43,124,190,.45)':'rgba(139,26,74,.4)'),
    borderColor:rows.map(r=>r.cmi>=1?BLUE:MAROON),borderWidth:1.3}]},
   options:{plugins:{legend:{display:false},tooltip:{callbacks:{
     label:c=>{const d=c.raw.d;return [tc(d.dept),'CMI '+d.cmi.toFixed(2)+' · '+d.dis+' cases',
       '₹'+(d.iprev/CR).toFixed(2)+' Cr · ALOS '+d.alos.toFixed(1)+' d'];}}}},
    scales:{x:{title:{display:true,text:'CMI'},min:0,
      grid:{color:ctx=>Math.abs(ctx.tick.value-1)<0.001?'#243342':'rgba(0,0,0,.06)'}},
     y:{title:{display:true,text:'discharges in month'},beginAtZero:true}},
    layout:{padding:24}}});

  // dept table
  document.getElementById('cmiTabNote').innerHTML='The <b>CMI</b> column is the specialty\'s fixed relative weight from the base period — '+
   'multiply it by that specialty\'s cases, sum across the table and divide by '+M.cmiN.toLocaleString('en-IN')+' cases to get this month\'s hospital CMI of <b>'+
   M.cmiHosp.toFixed(3)+'</b>. <b>₹/case</b> is the current month\'s actual, so a specialty billing above its own weight is running hotter than its base period. '+
   'Hospital average this month is <b>₹'+(M.avgC/L).toFixed(2)+' L per case</b> over '+M.TD+' discharges. '+
   'Intensity index does the same on a per-bed-day basis, so it rewards short-stay throughput where CMI does not. Click a header to sort.';
  const srt=rows.slice().sort((a,b)=>{const x=a[sortK],y=b[sortK];
   if(typeof x==='string')return sortA? x.localeCompare(y):y.localeCompare(x);
   return sortA? (x||0)-(y||0):(y||0)-(x||0);});
  const idxTag=v=>'<span class="tag '+(v>=1.15?'g':v<0.85?'r':'y')+'">'+v.toFixed(2)+'</span>';
  document.querySelector('#cmiTable tbody').innerHTML=srt.map(r=>
   '<tr'+(r.inBase?'':' style="opacity:.6"')+'><td class="doc">'+tc(r.dept)+
   (r.inBase?'':' <span class="tag y">day-care</span>')+'</td>'+
   '<td class="r">'+(r.iprev/CR).toFixed(2)+'</td><td class="r">'+r.dis+'</td>'+
   '<td class="r">'+r.alos.toFixed(2)+'</td><td class="r">'+(r.rpc/L).toFixed(2)+'</td>'+
   '<td class="r">'+idxTag(r.cmi)+'</td><td class="r">'+(r.rpb?(r.rpb/L).toFixed(3):'—')+'</td>'+
   '<td class="r">'+(r.ii?idxTag(r.ii):'—')+'</td>'+
   '<td class="r">'+(r.shRev*100).toFixed(1)+'%</td><td class="r">'+(r.shDis*100).toFixed(1)+'%</td>'+
   '<td class="r">'+(r.mom==null?'—':'<span class="tag '+(r.mom>=0?'g':'r')+'">'+(r.mom>=0?'+':'')+(r.mom/L).toFixed(2)+'</span>')+'</td>'+
   '<td class="r">'+r.docs+'</td></tr>').join('')+
   '<tr class="fytot"><td>All admitting specialties</td><td class="r">'+(M.TI/CR).toFixed(2)+'</td>'+
   '<td class="r">'+M.TD+'</td><td class="r">'+(M.TB/M.TD).toFixed(2)+'</td>'+
   '<td class="r">'+(M.avgC/L).toFixed(2)+'</td><td class="r">1.00</td>'+
   '<td class="r">'+(M.avgB/L).toFixed(3)+'</td><td class="r">1.00</td>'+
   '<td class="r">100%</td><td class="r">100%</td><td class="r"></td>'+
   '<td class="r">'+rows.reduce((a,r)=>a+r.docs,0)+'</td></tr>';

  // doctor table
  const q=(document.getElementById('cmiDocFilter').value||'').toLowerCase();
  const ds=M.docs.filter(d=>!q||d.doc.toLowerCase().includes(q)||d.dept.toLowerCase().includes(q))
   .sort((a,b)=>{const x=a[dSortK],y=b[dSortK];
    if(typeof x==='string')return dSortA? x.localeCompare(y):y.localeCompare(x);
    return dSortA? (x||0)-(y||0):(y||0)-(x||0);});
  document.querySelector('#cmiDocTable tbody').innerHTML=ds.map(d=>
   '<tr><td class="doc">'+tc(d.doc)+'</td><td class="doc">'+tc(d.dept)+'</td>'+
   '<td class="r">'+(d.iprev/L).toFixed(1)+'</td><td class="r">'+d.dis+'</td>'+
   '<td class="r">'+(d.alos?d.alos.toFixed(1):'—')+'</td><td class="r">'+(d.rpc/L).toFixed(2)+'</td>'+
   '<td class="r">'+idxTag(d.cmi)+'</td><td class="r">'+(d.rpb?(d.rpb/L).toFixed(3):'—')+'</td>'+
   '<td class="r">'+(d.ii?idxTag(d.ii):'—')+'</td>'+
   '<td class="r">'+(d.cash==null?'—':(d.cash*100).toFixed(0)+'%')+'</td></tr>').join('');

  // excluded
  document.getElementById('cmiExcl').innerHTML=M.exc.map(d=>
   '<span class="pill">'+tc(d.dept)+' · ₹'+(d.rev/CR).toFixed(2)+' Cr · '+d.dis+' disch</span>').join('')+
   '<div style="margin-top:8px;color:#7F8C9B;font-size:11.5px">Excluded revenue ₹'+
   (M.excRev/CR).toFixed(2)+' Cr of ₹'+((M.TI+M.excRev)/CR).toFixed(2)+
   ' Cr doctor-attributed. Gross revenue for the month per the flash is ₹'+(M.grossRev/CR).toFixed(2)+
   ' Cr — the gap is unallocated, non-doctor and OP-only revenue.</div>';

  // commentary
  const hiVol=rows.filter(r=>r.shDis>=0.04);
  const pull=hiVol.slice().sort((a,b)=>b.cmi-a.cmi)[0];
  const drag=hiVol.slice().sort((a,b)=>a.cmi-b.cmi)[0];
  const mism=rows.slice().sort((a,b)=>(b.ii-b.cmi)-(a.ii-a.cmi))[0];
  const mover=rows.filter(r=>r.mom!=null).sort((a,b)=>Math.abs(b.mom)-Math.abs(a.mom))[0];
  document.getElementById('cmiComment').innerHTML=
   'Hospital billed intensity is <b>₹'+(M.avgC/L).toFixed(2)+' L per inpatient case</b> at a blended ALOS of '+
   (M.TB/M.TD).toFixed(2)+' days'+(M.momAvg!=null? ', '+(M.momAvg>=0?'up':'down')+' '+
   Math.abs(M.momAvg*100).toFixed(1)+'% on '+mName(prevM):'')+'. '+
   (pull? 'Among specialties carrying at least 4% of cases, <b>'+tc(pull.dept)+'</b> lifts the mix hardest (CMI '+
    pull.cmi.toFixed(2)+' on '+pull.dis+' cases) while <b>'+tc(drag.dept)+'</b> dilutes it most (CMI '+
    drag.cmi.toFixed(2)+' on '+drag.dis+' cases). ':'')+
   (mism&&mism.ii-mism.cmi>0.3? '<b>'+tc(mism.dept)+'</b> is the clearest case of CMI understating economics — index '+
    mism.cmi.toFixed(2)+' per case but '+mism.ii.toFixed(2)+' per bed-day, i.e. short-stay throughput rather than low acuity. ':'')+
   (mover? 'Largest MoM move: <b>'+tc(mover.dept)+'</b> at '+(mover.mom>=0?'+':'')+'₹'+
    (mover.mom/L).toFixed(2)+' L per case. ':'')+
   '<span style="color:#7F8C9B">Caveat: billed intensity is not coded acuity. A specialty can score high because it uses expensive consumables or implants rather than because its patients are sicker, and pharmacy apportionment uses each doctor\'s IP revenue share as the key. Day-care-heavy specialties such as medical oncology and dialysis will read low on CMI and high on the intensity index by construction — read the two together, never CMI alone.</span>';
 }

 window.drawCmi=render;
 document.getElementById('cmiDocFilter').oninput=render;
 document.getElementById('cmiNoDay').onchange=render;
 document.getElementById('cmiMinBtns').innerHTML=[3,5,10].map(n=>
  '<button class="rbtn'+(n===minDoc?' on':'')+'" data-n="'+n+'">min '+n+' disch</button>').join('');
 document.querySelectorAll('#cmiMinBtns .rbtn').forEach(b=>b.onclick=()=>{
  minDoc=+b.dataset.n;
  document.querySelectorAll('#cmiMinBtns .rbtn').forEach(x=>x.classList.toggle('on',x===b));
  render();});
 document.querySelectorAll('#cmiTable thead th').forEach(t=>t.onclick=()=>{
  const k=t.dataset.k; if(!k)return; if(k===sortK)sortA=!sortA; else{sortK=k;sortA=false;} render();});
 document.querySelectorAll('#cmiDocTable thead th').forEach(t=>t.onclick=()=>{
  const k=t.dataset.k; if(!k)return; if(k===dSortK)dSortA=!dSortA; else{dSortK=k;dSortA=false;} render();});

 // headline card on the operations tab
 (function(){
  const M=model();
  const el=document.getElementById('cards');
  if(!el||!M.TD)return;
  const d=document.createElement('div');
  d.className='card m';
  const tp=M.rows.slice().sort((a,b)=>b.cmi-a.cmi)[0];
  const dc=(M.cmiPrev!=null&&M.cmiPrev)? M.cmiHosp-M.cmiPrev:null;
  d.innerHTML='<div class="lbl">Case Mix Index</div><div class="val">'+M.cmiHosp.toFixed(3)+'</div>'+
   '<div class="delta '+((dc==null||dc>=0)?'up':'dn')+'">₹'+(M.avgC/L).toFixed(2)+' L/case · '+
   (dc!=null? ((dc>=0?'▲ +':'▼ ')+dc.toFixed(3)+' vs '+mName(prevM)):'base month')+
   (tp? ' · top '+tc(tp.dept)+' '+tp.cmi.toFixed(2):'')+
   ' · <a href="#" onclick="showView(\'cmi\');return false;" style="color:#2B7CBE">detail</a></div>';
  el.appendChild(d);
 })();
})();

// ============================ FY27 FINANCIALS TAB ============================
(function(){
 const F=D.fy27;
 const tab=document.getElementById('tabFy27');
 if(!F){ if(tab) tab.style.display='none'; window.drawFy27=function(){}; return; }
 const MOS4=F.months||['Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar'];
 const units=Object.keys(F.units||{});
 let unit=units.includes('Kochi')?'Kochi':(units[0]||null);
 const charts={};
 const nn=v=>(typeof v==='number'&&isFinite(v));
 // last month index that actually carries data, so empty future months are dropped
 const lastIdx=(()=>{
  let li=-1;
  const probe=[(F.monthly['revenue:Kochi']||{}).total,(F.monthly.hospital||{}).ipDisch];
  probe.forEach(s=>{ if(s&&s.monthly) s.monthly.forEach((v,i)=>{ if(nn(v)&&Math.abs(v)>0) li=Math.max(li,i); }); });
  return li<0?3:li;
 })();
 const labs=MOS4.slice(0,lastIdx+1);
 const ser=(slot,key)=>{
  const b=F.monthly[slot]; if(!b||!b[key]) return [];
  return b[key].monthly.slice(0,lastIdx+1).map(v=>nn(v)?v:null);
 };

 document.getElementById('f27Through').textContent='through '+labs[labs.length-1]+'-'+
   String((new Date().getFullYear())%100).padStart(2,'0');
 document.getElementById('f27Basis').innerHTML=F.basisNote||'';
 document.getElementById('f27Gaps').innerHTML=
  'The <b>13Monthly P&amp;L</b> tab in both MIS packs evaluates to zero on every line for every month '+
  '(verified by recalculation, not a stale formula cache) and its columns are dated Apr-25&ndash;Mar-26, '+
  'so there is <b>no FY27 monthly EBITDA series</b> &mdash; EBITDA is shown for the report month and YTD only. '+
  'The <b>14KPI</b> tab has real volume counts but zero for every revenue-derived KPI and realization, '+
  'so <b>ARPOB and ALOS are not charted here</b>; use the Daily operations tab, which derives them from the '+
  'Daily MIS <i>MoM FY</i> sheet.';

 // ---- unit toggle ----
 const btnWrap=document.getElementById('f27UnitBtns');
 function paintBtns(){
  btnWrap.innerHTML=units.map(u=>'<button class="mbtn'+(u===unit?' on':'')+'" data-u="'+u+'">'+u+'</button>').join('');
  btnWrap.querySelectorAll('button').forEach(b=>b.onclick=()=>{unit=b.dataset.u;paintBtns();render();});
 }

 // ---- KPI cards + P&L table (per unit) ----
 function render(){
  const U=F.units[unit]; if(!U) return;
  const get=k=>U.lines.find(l=>l.key===k)||{};
  const rev=get('revenue'), eb=get('netEbitda').ytdAct!=null?get('netEbitda'):get('ebitda'), np=get('netProfit');
  const mar=(a,b)=>(nn(a)&&nn(b)&&b!==0)?(a/b*100):null;
  const ach=(a,b)=>(nn(a)&&nn(b)&&b!==0)?(a/b*100):null;
  const card=(lbl,val,sub,cls)=>'<div class="card'+(cls?' '+cls:'')+'"><div class="lbl">'+lbl+
    '</div><div class="val">'+val+'</div><div class="delta '+(sub&&sub.up?'up':'dn')+'">'+((sub&&sub.t)||'')+'</div></div>';
  const sgn=v=>v==null?'':(v>=0?'▲ ':'▼ ');
  const a1=ach(rev.ytdAct,rev.ytdBud);
  // EBITDA vs budget: a %-of-budget ratio is meaningless (and sign-flips) when the
  // budget is a loss, so show the absolute variance and colour on act>=bud instead.
  const ebVar=(nn(eb.ytdAct)&&nn(eb.ytdBud))? eb.ytdAct-eb.ytdBud : null;
  const ebPos=nn(eb.ytdBud)&&eb.ytdBud>0;
  const ebHead=ebVar==null? '—'
    : (ebPos? (eb.ytdAct/eb.ytdBud*100).toFixed(0)+'%'
            : (ebVar>=0?'+':'')+fmtCr(ebVar).replace('₹','₹'));
  document.getElementById('f27Cards').innerHTML=
   card('YTD revenue',fmtCr(rev.ytdAct||0),{t:(a1!=null? a1.toFixed(0)+'% of budget ('+fmtCr(rev.ytdBud||0)+')':''),up:(a1!=null&&a1>=100)})+
   card('YTD EBITDA',fmtCr(eb.ytdAct||0),{t:(mar(eb.ytdAct,rev.ytdAct)!=null? mar(eb.ytdAct,rev.ytdAct).toFixed(1)+'% margin':''),up:(eb.ytdAct||0)>=0},'m')+
   card('EBITDA vs budget',ebHead,
     {t:(nn(eb.ytdBud)? (ebPos? 'budget '+fmtCr(eb.ytdBud)
                              : (ebVar>=0?'better than':'worse than')+' budget '+fmtCr(eb.ytdBud)):''),
      up:(ebVar!=null&&ebVar>=0)},'m')+
   card(U.month+' revenue',fmtCr(rev.monthAct||0),{t:(ach(rev.monthAct,rev.monthBud)!=null? ach(rev.monthAct,rev.monthBud).toFixed(0)+'% of budget':''),up:(ach(rev.monthAct,rev.monthBud)||0)>=100})+
   card(U.month+' EBITDA',fmtCr(eb.monthAct||0),{t:(mar(eb.monthAct,rev.monthAct)!=null? mar(eb.monthAct,rev.monthAct).toFixed(1)+'% margin':''),up:(eb.monthAct||0)>=0},'m')+
   card('YTD net profit',fmtCr(np.ytdAct||0),{t:(mar(np.ytdAct,rev.ytdAct)!=null? mar(np.ytdAct,rev.ytdAct).toFixed(1)+'% of revenue':''),up:(np.ytdAct||0)>=0},'m');

  document.getElementById('f27PlNote').innerHTML='&#8377; Cr, '+unit+
   '. Actual vs budget for '+U.month+' and FY27 year-to-date. Budget is the MIS pack\'s own, which ties to the AOP basis.';
  const th=document.querySelector('#f27Pl thead'), tb=document.querySelector('#f27Pl tbody');
  th.innerHTML='<tr><th>Line</th><th class="r">'+U.month+' act</th><th class="r">'+U.month+' bud</th>'+
    '<th class="r">Var</th><th class="r">YTD act</th><th class="r">YTD bud</th><th class="r">Var</th><th class="r">% rev</th></tr>';
  const c=v=>nn(v)?(v/CR).toFixed(2):'—';
  const vcell=(a,b,favGood)=>{
   if(!nn(a)||!nn(b)) return '<td class="r flat">—</td>';
   const d=a-b, good=favGood? d>=0 : d<=0;
   return '<td class="r '+(Math.abs(d)<1e4?'flat':(good?'good':'bad'))+'">'+(d>=0?'+':'')+(d/CR).toFixed(2)+'</td>';
  };
  const COST=new Set(['consumables','staff','overheads','badDebts','totalExp','finance','depreciation']);
  const BOLD=new Set(['revenue','contribution','ebitda','netEbitda','cashProfit','netProfit']);
  const revYtd=rev.ytdAct;
  tb.innerHTML=U.lines.map(l=>{
   const favGood=!COST.has(l.key);
   const share=(nn(l.ytdAct)&&nn(revYtd)&&revYtd)?(l.ytdAct/revYtd*100).toFixed(1)+'%':'—';
   return '<tr'+(BOLD.has(l.key)?' class="fytot"':'')+'><td>'+l.label+'</td>'+
    '<td class="r">'+c(l.monthAct)+'</td><td class="r">'+c(l.monthBud)+'</td>'+vcell(l.monthAct,l.monthBud,favGood)+
    '<td class="r">'+c(l.ytdAct)+'</td><td class="r">'+c(l.ytdBud)+'</td>'+vcell(l.ytdAct,l.ytdBud,favGood)+
    '<td class="r">'+share+'</td></tr>';
  }).join('');

  drawRev(); drawOpv();
 }

 const mk=(id,cfg)=>{ const el=document.getElementById(id); if(!el) return;
  if(charts[id]) charts[id].destroy(); charts[id]=new Chart(el,cfg); };
 const money={ticks:{callback:v=>(v/CR).toFixed(0)}};

 function drawRev(){
  const slot='revenue:'+unit;
  mk('f27RevChart',{type:'bar',data:{labels:labs,datasets:[
    {label:'OP',data:ser(slot,'op'),backgroundColor:BLUE,stack:'s'},
    {label:'IP',data:ser(slot,'ip'),backgroundColor:MAROON,stack:'s'}]},
   options:{responsive:true,plugins:{tooltip:{callbacks:{label:c=>c.dataset.label+': '+fmtCr(c.parsed.y||0)}}},
    scales:{x:{stacked:true},y:{stacked:true,...money,title:{display:true,text:'₹ Cr'}}}}});
 }
 function drawOpv(){
  const slot='opVisits:'+unit;
  mk('f27OpvChart',{type:'bar',data:{labels:labs,datasets:[
    {label:'New',data:ser(slot,'new'),backgroundColor:MAROON,stack:'v'},
    {label:'Revisit',data:ser(slot,'revisit'),backgroundColor:LT,stack:'v'}]},
   options:{responsive:true,scales:{x:{stacked:true},y:{stacked:true,title:{display:true,text:'visits'}}}}});
 }
 function drawGroup(){
  const h=F.monthly.hospital||{};
  const occ=(h.occupancy?h.occupancy.monthly.slice(0,lastIdx+1):[]).map(v=>nn(v)?v*100:null);
  mk('f27OccChart',{data:{labels:labs,datasets:[
    {type:'bar',label:'IP discharges',data:(h.ipDisch?h.ipDisch.monthly.slice(0,lastIdx+1):[]).map(v=>nn(v)?v:null),backgroundColor:BLUE,yAxisID:'y'},
    {type:'line',label:'Occupancy %',data:occ,borderColor:MAROON,backgroundColor:MAROON,tension:.3,yAxisID:'y1'}]},
   options:{responsive:true,scales:{y:{title:{display:true,text:'discharges'}},
    y1:{position:'right',grid:{drawOnChartArea:false},title:{display:true,text:'occupancy %'},
     ticks:{callback:v=>v.toFixed(0)+'%'}}}}});
  const A=F.monthly.payorA||{}, B=F.monthly.payorB||{};
  const pk=[['cash','Cash',BLUE],['insurance','Insurance',MAROON],['government','Government',GOLD],
            ['corporate','Corporate',GRAY],['international','International',LT]];
  mk('f27PayChart',{type:'bar',data:{labels:labs,datasets:pk.map(([k,lbl,col])=>{
    const src=(A[k]?A:(B[k]?B:null));
    return {label:lbl,data:src?src[k].monthly.slice(0,lastIdx+1).map(v=>nn(v)?v:null):[],backgroundColor:col,stack:'p'};
   })},
   options:{responsive:true,plugins:{tooltip:{callbacks:{label:c=>c.dataset.label+': '+fmtCr(c.parsed.y||0)}}},
    scales:{x:{stacked:true},y:{stacked:true,...money,title:{display:true,text:'₹ Cr'}}}}});
 }

 // ---------------- AOP tracker panels ----------------
 const T=F.tracker;
 let trkMode='var';
 window.trkSort=function(m){ trkMode=m;
  document.getElementById('trkSortVar').classList.toggle('on',m==='var');
  document.getElementById('trkSortOrd').classList.toggle('on',m==='ord');
  renderTrkPl(); };

 function trkLine(lbl){ return (T.lines||[]).find(l=>l.label===lbl); }

 function renderTracker(){
  if(!T){ const w=document.getElementById('f27TrackerWrap'); if(w) w.style.display='none'; return; }
  document.getElementById('trkSrc').textContent=T.source||'';
  // how many months carry actuals
  const ch=T.chart||{labels:[],revenue:[],ebitda:[]};
  const nAct=ch.revenue.filter(x=>nn(x.act)).length;
  const rev=trkLine('Total Revenue')||{}, eb=trkLine('EBITDA')||{}, pat=trkLine('PAT')||{};
  const fyPlan=rev.fy27Plan;
  document.getElementById('trkNote').innerHTML=
   'The AOP plan workbook, in &#8377; Lakhs at source. Its FY27 revenue plan of <b>'+fmtCr(fyPlan||0)+
   '</b> ties exactly to the AOP driving the daily budget lines, and plan-total less IP+OP equals '+
   'F&amp;B plus Other Income to the rupee &mdash; so this is the same plan, not a third basis. '+
   'Actuals cover <b>'+nAct+' month'+(nAct===1?'':'s')+'</b> (Apr&ndash;'+(ch.labels[nAct-1]||'')+').';

  const ach=(a,b)=>(nn(a)&&nn(b)&&b!==0)? a/b*100 : null;
  const vs=(a,b,favGood)=>{ const r=ach(a,b); if(r==null) return {t:'',up:true};
   const good=favGood? r>=100 : r<=100;
   return {t:r.toFixed(0)+'% of plan ('+fmtCr(b)+')',up:good}; };
  const st=(lbl,val,sub,cls)=>'<div class="vstat"><div class="vlbl">'+lbl+'</div><div class="vval'+
    (cls?' '+cls:'')+'">'+val+'</div><div class="vsub">'+sub+'</div></div>';
  const mg=(a,b)=>(nn(a)&&nn(b)&&b)? (a/b*100).toFixed(1)+'%' : '—';
  document.getElementById('trkStrip').innerHTML=
   st('YTD revenue',fmtCr(rev.ytdAct||0),vs(rev.ytdAct,rev.ytdPlan,true).t,
      (ach(rev.ytdAct,rev.ytdPlan)>=100?'good':'bad'))+
   st('YTD EBITDA',fmtCr(eb.ytdAct||0),'margin '+mg(eb.ytdAct,rev.ytdAct)+' · plan '+mg(eb.ytdPlan,rev.ytdPlan),
      (ach(eb.ytdAct,eb.ytdPlan)>=100?'good':'bad'))+
   st('YTD PAT',fmtCr(pat.ytdAct||0),'margin '+mg(pat.ytdAct,rev.ytdAct),
      (ach(pat.ytdAct,pat.ytdPlan)>=100?'good':'bad'))+
   st('FY27 plan',fmtCr(fyPlan||0),'FY26 actual '+fmtCr(rev.fy26Act||0)+' · +'+
      ((rev.fy26Act&&fyPlan)?((fyPlan/rev.fy26Act-1)*100).toFixed(0):'—')+'%');

  // dual monthly plan note — pinned to the latest month the tracker actually carries
  const M3={Apr:'April',May:'May',Jun:'June',Jul:'July',Aug:'August',Sep:'September',
    Oct:'October',Nov:'November',Dec:'December',Jan:'January',Feb:'February',Mar:'March'};
  const want=window.__closedMonth||null;   // the month the rest of the dashboard reports on
  let augIdx=want? ch.labels.findIndex(l=>M3[l]===want) : -1;
  if(augIdx<0||!nn((ch.revenue[augIdx]||{}).plan))
    augIdx=(function(){for(let i=ch.labels.length-1;i>=0;i--){
      const r=ch.revenue[i]; if(r&&nn(r.plan)&&nn(r.act)&&r.act>0) return i;} return -1;})();
  const trkAug=augIdx>=0? ch.revenue[augIdx].plan : null;
  const augLbl=augIdx>=0? ch.labels[augIdx] : '—';
  const augFull=M3[augLbl]||augLbl;
  const aopAug=(D.aop&&D.aop.months)? (D.aop.months.find(m=>m.month===augFull)||{}).total : null;
  document.getElementById('trkAugMon').textContent=augFull;
  // Which phasing do the daily budget lines actually follow? Decide it from the
  // data rather than asserting it — the two plans differ by month.
  const dailyBudMon=(function(){
    const mi=Object.keys(D.daily).find(k=>{
      const [y,m]=k.split('-').map(Number);
      return ({April:4,May:5,June:6,July:7,August:8,September:9,October:10,
               November:11,December:12,January:1,February:2,March:3})[augFull]===m;});
    return mi? D.daily[mi].reduce((a,r)=>a+(r.budTot||0),0):null;})();
  const near=(a,b)=>nn(a)&&nn(b)&&Math.abs(a-b)<Math.max(1e5,Math.abs(b)*0.002);
  let tiesTo=null;
  if(near(dailyBudMon,trkAug)&&!near(dailyBudMon,aopAug)) tiesTo='tracker';
  else if(near(dailyBudMon,aopAug)&&!near(dailyBudMon,trkAug)) tiesTo='aop';
  document.getElementById('trkAugNote').innerHTML=
   'The same annual plan is phased two ways. The AOP monthly phasing puts '+augFull+' at <b>'+
   (nn(aopAug)?fmtCr(aopAug):'—')+'</b>; this tracker puts it at <b>'+
   (nn(trkAug)?fmtCr(trkAug):'—')+'</b>'+
   ((nn(trkAug)&&nn(aopAug))? ' &mdash; a difference of '+fmtCr(Math.abs(trkAug-aopAug))+' ('+
     (Math.abs(trkAug/aopAug-1)*100).toFixed(1)+'%)':'')+
   '. The annual totals are identical; only the month-by-month split differs. '+
   (nn(dailyBudMon)
     ? 'The daily budget lines on the <b>Daily operations</b> tab sum to <b>'+fmtCr(dailyBudMon)+
       '</b> for '+augFull+
       (tiesTo==='tracker'
         ? ', which ties to the <b>tracker phasing</b> &mdash; so the achievement percentages '+
           'shown elsewhere on this dashboard are measured against that split, not the AOP one.'
         : tiesTo==='aop'
         ? ', which ties to the <b>AOP phasing</b> &mdash; so the achievement percentages shown '+
           'elsewhere on this dashboard are measured against the AOP split.'
         : ', which matches neither phasing exactly; treat the achievement percentages as '+
           'indicative for this month and confirm the budget source with Finance.')
     : 'No daily budget series is available for '+augFull+' to reconcile against.');

  renderTrkPl(); renderTrkInit(); drawTrkEb();
 }

 function renderTrkPl(){
  const rev=trkLine('Total Revenue')||{};
  document.getElementById('trkPlNote').innerHTML='&#8377; Cr, FY27 year-to-date. '+
   'Variance is favourable when revenue beats plan or a cost line comes in under it &mdash; '+
   'cost lines are scored in that direction, so green always means good.';
  let rows=(T.lines||[]).filter(l=>nn(l.ytdPlan)||nn(l.ytdAct));
  const varPc=l=>{ if(!nn(l.ytdPlan)||!l.ytdPlan||!nn(l.ytdAct)) return 0;
   const d=(l.ytdAct-l.ytdPlan)/Math.abs(l.ytdPlan)*100;
   return l.isCost? -d : d; };
  if(trkMode==='var') rows=rows.slice().sort((a,b)=>varPc(a)-varPc(b));
  const th=document.querySelector('#trkPl thead'), tb=document.querySelector('#trkPl tbody');
  th.innerHTML='<tr><th>P&amp;L line</th><th class="r">FY26 act</th><th class="r">FY27 plan</th>'+
   '<th class="r">YTD plan</th><th class="r">YTD act</th><th class="r">Var</th><th class="r">% of plan</th></tr>';
  const c=v=>nn(v)?(v/CR).toFixed(2):'—';
  tb.innerHTML=rows.map(l=>{
   const d=(nn(l.ytdAct)&&nn(l.ytdPlan))? l.ytdAct-l.ytdPlan : null;
   const fav=(d==null)? null : (l.isCost? d<=0 : d>=0);
   const r=(nn(l.ytdPlan)&&l.ytdPlan)? (l.ytdAct/l.ytdPlan*100) : null;
   return '<tr'+(l.isSub?' class="fytot"':'')+'><td>'+l.label+'</td>'+
    '<td class="r">'+c(l.fy26Act)+'</td><td class="r">'+c(l.fy27Plan)+'</td>'+
    '<td class="r">'+c(l.ytdPlan)+'</td><td class="r">'+c(l.ytdAct)+'</td>'+
    '<td class="r '+(d==null?'flat':(fav?'good':'bad'))+'">'+(d==null?'—':((d>=0?'+':'')+(d/CR).toFixed(2)))+'</td>'+
    '<td class="r '+(r==null?'flat':(fav?'good':'bad'))+'">'+(r==null?'—':r.toFixed(0)+'%')+'</td></tr>';
  }).join('');
 }

 function renderTrkInit(){
  const I=T.initiatives||[], tt=T.initTotals||{};
  const rev=trkLine('Total Revenue')||{};
  const growth=(nn(rev.fy27Plan)&&nn(rev.fy26Act))? rev.fy27Plan-rev.fy26Act : null;
  document.getElementById('trkInitNote').innerHTML='&#8377; Cr. Full-year target and the phased build-up to date. '+
   'Revenue initiatives total <b>'+fmtCr((tt.revImpact||{}).fullYear||0)+'</b> and cost savings <b>'+
   fmtCr((tt.costSaving||{}).fullYear||0)+'</b>, a gross <b>'+fmtCr((tt.gross||{}).fullYear||0)+'</b>'+
   (nn(tt.extraCost&&tt.extraCost.total)? ', against <b>'+fmtCr(tt.extraCost.total)+
     '</b> of additional cost needed to land the revenue uplift':'')+'. '+
   (growth!=null? 'For scale, the plan lifts revenue '+fmtCr(growth)+' over FY26 actual, so the revenue '+
     'initiatives account for essentially all of the planned growth.':'');
  const th=document.querySelector('#trkInit thead'), tb=document.querySelector('#trkInit tbody');
  const nMonths=12;
  th.innerHTML='<tr><th>Initiative</th><th>Type</th><th class="r">Full year</th>'+
   '<th class="r">Booked to date</th><th class="r">% of target</th></tr>';
  // "booked to date" = sum of phased months that have elapsed (same count as actuals)
  const ch=T.chart||{revenue:[]};
  const nAct=Math.max(1,ch.revenue.filter(x=>nn(x.act)).length);
  let html='', section=null;
  I.forEach(it=>{
   if(it.section!==section){ section=it.section;
    html+='<tr class="sep"><td colspan="5" style="font-weight:700;color:#8B1A4A">'+(section||'')+'</td></tr>'; }
   const booked=it.months.slice(0,nAct).reduce((a,v)=>a+(nn(v)?v:0),0);
   const pc=(nn(it.fullYear)&&it.fullYear)? booked/it.fullYear*100 : null;
   html+='<tr><td>'+it.name+'</td><td style="color:#7F8C9B">'+(it.type||'')+'</td>'+
    '<td class="r">'+(nn(it.fullYear)?(it.fullYear/CR).toFixed(2):'—')+'</td>'+
    '<td class="r">'+(booked/CR).toFixed(2)+'</td>'+
    '<td class="r '+(pc==null?'flat':(pc>=nAct/nMonths*100?'good':'bad'))+'">'+
      (pc==null?'—':pc.toFixed(0)+'%')+'</td></tr>';
  });
  tb.innerHTML=html;
 }

 function drawTrkEb(){
  const ch=T.chart; if(!ch) return;
  const rev=ch.revenue.map(x=>x.act), eb=ch.ebitda;
  const marg=eb.map((x,i)=>(nn(x.act)&&nn(rev[i])&&rev[i])? x.act/rev[i]*100 : null);
  mk('trkEbChart',{data:{labels:ch.labels,datasets:[
    {type:'bar',label:'EBITDA actual',data:eb.map(x=>nn(x.act)?x.act:null),backgroundColor:BLUE,yAxisID:'y'},
    {type:'line',label:'EBITDA plan',data:eb.map(x=>nn(x.plan)?x.plan:null),borderColor:MAROON,
     backgroundColor:MAROON,borderDash:[5,4],tension:.25,yAxisID:'y'},
    {type:'line',label:'Margin % (actual)',data:marg,borderColor:GOLD,backgroundColor:GOLD,
     tension:.25,yAxisID:'y1',spanGaps:false}]},
   options:{responsive:true,
    plugins:{tooltip:{callbacks:{label:c=>c.dataset.label+': '+
      (c.dataset.yAxisID==='y1'? (c.parsed.y==null?'—':c.parsed.y.toFixed(1)+'%') : fmtCr(c.parsed.y||0))}}},
    scales:{y:{...money,title:{display:true,text:'₹ Cr'}},
     y1:{position:'right',grid:{drawOnChartArea:false},title:{display:true,text:'margin %'},
      ticks:{callback:v=>v.toFixed(0)+'%'}}}}});
 }

 let drawn=false;
 window.drawFy27=function(){ paintBtns(); render(); renderTracker();
  if(!drawn){drawGroup();drawn=true;} };
})();
</script></body></html>
"""

if __name__ == "__main__":
    main()
